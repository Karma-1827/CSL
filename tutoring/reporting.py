from collections import defaultdict
from decimal import Decimal
from io import BytesIO, StringIO
import re
from xml.sax.saxutils import escape

from django.core.exceptions import ValidationError
from django.db.models import Q, Sum
from django.conf import settings
from django.utils import timezone

from accounts.models import EducationLevel, PartnerProgram, Role

from .models import ClassSession, ClassSessionStatus, HourAdjustment
from .services import LEVEL_LABELS, class_is_valid


def user_has_hour_records(user):
    if user.role == Role.TUTOR:
        return True
    return bool(
        user.role == Role.TUTEE
        and user.roster_entry
        and user.roster_entry.program_id
        and user.roster_entry.program.tutee_can_download_hours
    )


def tutor_available_programs(tutor):
    """Distinct partner programs among a tutor's tutees or manual hour adjustments,
    restricted to those with a configured tutor-facing certificate template."""
    session_program_ids = ClassSession.objects.filter(pairing__tutor=tutor).values_list(
        "pairing__tutee__roster_entry__program_id", flat=True
    )
    adjustment_program_ids = HourAdjustment.objects.filter(user=tutor).values_list("program_id", flat=True)
    program_ids = set(session_program_ids) | set(adjustment_program_ids)
    return PartnerProgram.objects.filter(
        pk__in=program_ids
    ).exclude(tutor_certificate_filename="").order_by("name_zh")


def hour_adjustment_total(user, starts_on, ends_on, program=None):
    """Sum manually-credited hours whose semester fully falls within [starts_on, ends_on]."""
    rows = HourAdjustment.objects.filter(
        user=user, semester__starts_on__gte=starts_on, semester__ends_on__lte=ends_on,
    )
    if program is not None:
        rows = rows.filter(program=program)
    return rows.aggregate(total=Sum("hours"))["total"] or Decimal("0")


def valid_sessions_for_user(user, starts_on, ends_on, program=None):
    participant = Q(pairing__tutor=user) if user.role == Role.TUTOR else Q(pairing__tutee=user)
    rows = ClassSession.objects.filter(
        participant,
        status=ClassSessionStatus.SCHEDULED,
        class_date__range=(starts_on, ends_on),
    ).select_related(
        "pairing__semester", "pairing__tutor", "pairing__tutee", "pairing__tutee__tutee_profile"
    ).prefetch_related(
        "attendances", "class_records", "confirmations", "makeup_review"
    ).order_by("class_date", "start_time")
    if program is not None:
        rows = rows.filter(pairing__tutee__roster_entry__program=program)
    return [row for row in rows if class_is_valid(row)]


def hour_report_data(user, starts_on, ends_on, program=None):
    sessions = valid_sessions_for_user(user, starts_on, ends_on, program=program)
    sections = defaultdict(list)
    for session in sessions:
        own_record = next((r for r in session.class_records.all() if r.author_id == user.pk), None)
        counterpart = session.pairing.tutee if user.role == Role.TUTOR else session.pairing.tutor
        student_profile = getattr(session.pairing.tutee, "tutee_profile", None)
        sections[session.pairing.semester].append({
            "session": session,
            "counterpart": counterpart,
            "topic": own_record.topic if own_record else "",
            "student_nationality": student_profile.nationality if student_profile else "—",
            "student_level": LEVEL_LABELS.get(student_profile.overall_level, student_profile.overall_level) if student_profile else "—",
        })
    result = []
    for semester, rows in sorted(sections.items(), key=lambda item: item[0].starts_on):
        result.append({"semester": semester, "rows": rows, "subtotal": sum(r["session"].duration for r in rows)})
    session_total = sum(row.duration for row in sessions)
    adjustment_total = hour_adjustment_total(user, starts_on, ends_on, program=program)
    return {
        "user": user,
        "starts_on": starts_on,
        "ends_on": ends_on,
        "sections": result,
        "session_total": session_total,
        "adjustment_total": adjustment_total,
        "total": session_total + adjustment_total,
        "generated_at": timezone.localtime(),
    }


_CERTIFICATE_FONT_NAME = "CertificateLiSong"
_CERTIFICATE_BOLD_FONT_NAME = "CertificateLiSong-Bold"
_CERTIFICATE_ENGLISH_FONT_NAME = "CertificateHelveticaNeue"
_CERTIFICATE_ENGLISH_BOLD_FONT_NAME = "CertificateHelveticaNeue-Bold"


def _register_certificate_fonts():
    """Register the shared CJK/Latin font pair used by every PDF this module produces.

    Returns (font_name, bold_font_name, english_font_name, english_bold_font_name).
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = _CERTIFICATE_FONT_NAME
    bold_font_name = _CERTIFICATE_BOLD_FONT_NAME
    english_font_name = _CERTIFICATE_ENGLISH_FONT_NAME
    english_bold_font_name = _CERTIFICATE_ENGLISH_BOLD_FONT_NAME
    font_dir = settings.BASE_DIR / "assets/fonts"
    chinese_regular = font_dir / "DFLiSongStd-W3.ttf"
    chinese_bold = font_dir / "DFLiSongStd-W7.ttf"
    english_font = font_dir / "Helvetica Neue Condensed Bold.ttf"
    # ReportLab cannot embed the CFF/PostScript outlines in the department's
    # original OTF files. The private deployment therefore provides converted
    # TrueType copies beside them. Open-licensed fallbacks keep source-only
    # installations and CI functional when those private assets are absent.
    if not chinese_regular.exists():
        chinese_regular = font_dir / "TW-Kai.ttf"
    if not chinese_bold.exists():
        chinese_bold = chinese_regular
    if english_font.exists():
        english_regular = english_font
        english_bold = english_font
    else:
        english_regular = font_dir / "LiberationSerif-Regular.ttf"
        english_bold = font_dir / "LiberationSerif-Bold.ttf"
    if font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(font_name, chinese_regular))
    if bold_font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(bold_font_name, chinese_bold))
    if english_font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(english_font_name, english_regular))
    if english_bold_font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(english_bold_font_name, english_bold))
    pdfmetrics.registerFontFamily(
        font_name,
        normal=font_name,
        bold=bold_font_name,
        italic=font_name,
        boldItalic=bold_font_name,
    )
    pdfmetrics.registerFontFamily(
        english_font_name,
        normal=english_font_name,
        bold=english_bold_font_name,
        italic=english_font_name,
        boldItalic=english_bold_font_name,
    )
    return font_name, bold_font_name, english_font_name, english_bold_font_name


def _restrict_copy_and_selection(pdf_bytes):
    """Best-effort "no copy / no text extraction" protection for every PDF this module
    produces (item 3 of MEETING_CHANGE_REQUIREMENTS_2026-08-04.md).

    This only sets a PDF permission flag honored by compliant readers — it is NOT real DRM.
    It cannot stop screenshots, OCR, re-typing, or tools that deliberately ignore permission
    flags (see pypdf's own docs; this is a property of the PDF encryption spec, not a pypdf
    limitation). No user/open password is set, so the file still opens directly; printing is
    still granted. The random owner password is generated per file and discarded immediately
    — nothing in this codebase ever needs to remove the restriction again.
    """
    from pypdf import PdfReader, PdfWriter
    from pypdf.constants import UserAccessPermissions
    import secrets

    reader = PdfReader(BytesIO(pdf_bytes))
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    writer.encrypt(
        user_password="",
        owner_password=secrets.token_urlsafe(24),
        permissions_flag=UserAccessPermissions.PRINT | UserAccessPermissions.PRINT_TO_REPRESENTATION,
    )
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def build_hours_pdf(data, *, version="summary", detail_fields=(), program=None, language="zh"):
    """Overlay a formal certificate onto the department-provided PDF template.

    `language` selects a monolingual certificate ("zh" or "en") — see
    MEETING_CHANGE_REQUIREMENTS_2026-08-04.md item 13. The one exception is the user's own name,
    which always follows display_name_markup()'s rule regardless of certificate language.
    """
    from pypdf import PdfReader, PdfWriter
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas
    from reportlab.platypus import Paragraph, Table, TableStyle

    if language not in {"zh", "en"}:
        raise ValidationError("證明語言不正確。 / Invalid certificate language.")
    is_zh = language == "zh"

    font_name, bold_font_name, english_font_name, english_bold_font_name = _register_certificate_fonts()

    user = data["user"]
    roster = user.roster_entry
    if user.role == Role.TUTEE:
        program = roster.program if roster else None
    if program is None:
        raise ValidationError("此帳號尚未設定合作計畫，無法產生證明。 / No partner program is configured for this account.")
    if user.role == Role.TUTEE:
        template_name = program.tutee_certificate_filename
        title_zh = program.tutee_certificate_title_zh
        title_en = program.tutee_certificate_title_en
        plan_name = program.tutee_certificate_plan_name
        plan_name_en = program.tutee_certificate_plan_name_en
        activity = program.tutee_certificate_activity_text
        activity_en = program.tutee_certificate_activity_text_en
    else:
        template_name = program.tutor_certificate_filename
        title_zh = program.tutor_certificate_title_zh
        title_en = program.tutor_certificate_title_en
        plan_name = program.tutor_certificate_plan_name
        plan_name_en = program.tutor_certificate_plan_name_en
        activity = program.tutor_certificate_activity_text
        activity_en = program.tutor_certificate_activity_text_en
    if not template_name:
        raise ValidationError(
            "此合作計畫尚未設定證明模板，請聯絡系辦。 / This partner program has no certificate template configured."
        )
    is_ntnu_tutor = program.code == "NTNU" and user.role == Role.TUTOR
    title = title_zh if is_zh else title_en
    # The NTNU-tutor branch below has its own hardcoded body text (a specific fixed-format
    # sentence required by the department) and never reads plan_name/activity, so only its
    # title needs to exist; every other program×role combination renders plan_name/activity
    # into the body and must have both set for whichever language was requested (item 16:
    # any missing title/plan_name/activity, in either language, must fail clearly here
    # rather than render a certificate with a blank clause).
    if is_ntnu_tutor:
        missing_content = False
    else:
        plan_name_value = plan_name if is_zh else plan_name_en
        activity_value = activity if is_zh else activity_en
        missing_content = not plan_name_value or not activity_value
    if not title or missing_content:
        raise ValidationError(
            "此合作計畫尚未設定所選語言的證明文案，請洽系辦設定。 / "
            "This partner program has no certificate text configured for the selected language yet."
        )
    template_path = settings.BASE_DIR / "tutoring/resources/certificate_templates" / template_name

    def hours_text(value):
        value = str(value)
        return value[:-2] if value.endswith(".0") else value

    def roc_date(value):
        return f"中華民國 {value.year - 1911} 年 {value.month} 月 {value.day} 日"

    def gregorian_date(value):
        return f"{value.strftime('%B')} {value.day}, {value.year}"

    def period_markup():
        start, end = data["starts_on"], data["ends_on"]
        bold_number = lambda value: f'<font name="{english_bold_font_name}">{value}</font>'
        start_text = (
            f"民國{bold_number(start.year - 1911)}年"
            f"{bold_number(start.month)}月{bold_number(start.day)}日"
        )
        if start.year == end.year:
            end_text = f"{bold_number(end.month)}月{bold_number(end.day)}日"
        else:
            end_text = (
                f"{bold_number(end.year - 1911)}年"
                f"{bold_number(end.month)}月{bold_number(end.day)}日"
            )
        return f"{start_text}至{end_text}"

    def period_markup_en():
        start, end = data["starts_on"], data["ends_on"]
        return (
            f"{start.strftime('%B')} {start.day}, {start.year} to "
            f"{end.strftime('%B')} {end.day}, {end.year}"
        )

    def month_period_markup():
        start, end = data["starts_on"], data["ends_on"]
        start_year = start.year - 1911
        end_year = end.year - 1911
        bold_number = lambda value: f'<font name="{english_bold_font_name}">{value}</font>'
        if start.year == end.year and start.month == end.month:
            return f"民國{bold_number(start_year)}年{bold_number(start.month)}月"
        if start.year == end.year:
            return (
                f"民國{bold_number(start_year)}年"
                f"{bold_number(start.month)}月-{bold_number(end.month)}月"
            )
        return (
            f"民國{bold_number(start_year)}年{bold_number(start.month)}月-"
            f"{bold_number(end_year)}年{bold_number(end.month)}月"
        )

    def month_period_markup_en():
        start, end = data["starts_on"], data["ends_on"]
        if start.year == end.year and start.month == end.month:
            return f"{start.strftime('%B')} {start.year}"
        if start.year == end.year:
            return f"{start.strftime('%B')}–{end.strftime('%B')} {start.year}"
        return f"{start.strftime('%B')} {start.year} – {end.strftime('%B')} {end.year}"

    latin_run_pattern = re.compile(r"[A-Za-z0-9][A-Za-z0-9 .,/():+\-]*")

    def mixed_font_markup(value, *, bold=False):
        """Use Helvetica Neue for Latin text while retaining LiSong for Chinese.

        Every run (Latin and CJK alike) gets an explicit <font name="..."> tag rather
        than relying on the surrounding Paragraph's default font or on <b> resolving
        through registerFontFamily(): the English certificate paragraphs default to
        the English certificate family, whose registered "bold" face is Latin-only, so a bare <b>
        around Chinese text there silently drops the glyphs instead of rendering them.
        """
        value = str(value)
        result = []
        cursor = 0
        latin_font = english_bold_font_name if bold else english_font_name
        cjk_font = bold_font_name if bold else font_name
        for match in latin_run_pattern.finditer(value):
            if match.start() > cursor:
                result.append(f'<font name="{cjk_font}">{escape(value[cursor:match.start()])}</font>')
            result.append(f'<font name="{latin_font}">{escape(match.group())}</font>')
            cursor = match.end()
        if cursor < len(value):
            result.append(f'<font name="{cjk_font}">{escape(value[cursor:])}</font>')
        return "".join(result)

    def display_name_markup():
        """Name display is the one exception to the single-language rule (item 13): both
        names present always show "中文姓名 / English Name" regardless of certificate
        language; only one present shows just that one, with no dangling slash.

        Font names are set explicitly (not via <b>) for the same reason documented in
        mixed_font_markup(): <b> resolution depends on the surrounding paragraph's
        default font family, which differs between the Chinese and English certificates.
        """
        if user.name_zh and user.name_en:
            return (
                f'<font name="{bold_font_name}">{escape(user.name_zh)}</font> / '
                f'<font name="{english_bold_font_name}">{escape(user.name_en)}</font>'
            )
        return mixed_font_markup(user.bilingual_name, bold=True)

    certificate_lead = None
    if is_ntnu_tutor:
        education_level_labels_zh = {
            EducationLevel.BACHELOR: "大學部",
            EducationLevel.MASTER: "碩士班",
            EducationLevel.DOCTORAL: "博士班",
        }
        education_level_labels_en = {
            EducationLevel.BACHELOR: "undergraduate",
            EducationLevel.MASTER: "graduate (Master's)",
            EducationLevel.DOCTORAL: "doctoral",
        }
        if is_zh:
            level_label = education_level_labels_zh.get(roster.education_level, "") if roster else ""
            certificate_lead = (
                f"本系{level_label}學生<br/>"
                f"{display_name_markup()}（學號："
                f'<font name="{english_bold_font_name}">{escape(user.username)}</font>）'
            )
            certificate_paragraph = (
                f"於{month_period_markup()}，"
                f"於本校擔任國際生華語輔導老師，總計授課 "
                f'<font name="{english_bold_font_name}">{hours_text(data["total"])}</font>'
                f"<b> 小時</b>。特此證明"
            )
        else:
            level_label = education_level_labels_en.get(roster.education_level, "") if roster else ""
            certificate_lead = (
                f'<font name="{english_font_name}">This certifies that {level_label} student</font><br/>'
                f"{display_name_markup()} "
                f'<font name="{english_font_name}">(Student ID: {escape(user.username)})</font>'
            )
            certificate_paragraph = (
                f'<font name="{english_font_name}">served as a Chinese tutor for international '
                f"students at National Taiwan Normal University during {month_period_markup_en()}, "
                f'completing a total of <font name="{english_bold_font_name}">{hours_text(data["total"])}</font>'
                f' teaching hours. This certifies the above.</font>'
            )
        summary_paragraph_style = ParagraphStyle(
            "CertificateSummaryBodyNtnuTutor", fontName=font_name if is_zh else english_font_name,
            fontSize=20 if is_zh else 15, leading=43 if is_zh else 26,
            alignment=TA_JUSTIFY, firstLineIndent=0,
            textColor=colors.HexColor("#151515"), wordWrap="CJK" if is_zh else None,
        )
    else:
        display_name = display_name_markup()
        if is_zh:
            certificate_lead = (
                f"茲證明<br/>{display_name} 同學（學號："
                f'<font name="{english_bold_font_name}">{escape(user.username)}</font>）'
            )
            certificate_paragraph = (
                f"於<b>{period_markup()}</b>期間，參與國立臺灣師範大學華語文教學系"
                f"「<b>{plan_name}</b>」，{activity}共計 "
                f'<font name="{english_bold_font_name}">{hours_text(data["total"])}</font>'
                f"<b> 小時</b>，特此證明。"
            )
        else:
            certificate_lead = (
                f'<font name="{english_font_name}">This is to certify that</font><br/>'
                f'{display_name} <font name="{english_font_name}">'
                f'(Student ID: {escape(user.username)})</font>'
            )
            certificate_paragraph = (
                f'<font name="{english_font_name}">participated in the '
                f'&#8220;<b>{escape(plan_name_en)}</b>&#8221; of the Department of Chinese as a Second '
                f"Language, National Taiwan Normal University, during <b>{period_markup_en()}</b>, "
                f'{escape(activity_en)}, completing a total of '
                f'<font name="{english_bold_font_name}">{hours_text(data["total"])}</font> hours. '
                f"This certificate is issued accordingly.</font>"
            )
        summary_paragraph_style = ParagraphStyle(
            "CertificateSummaryBody", fontName=font_name if is_zh else english_font_name,
            fontSize=15 if is_zh else 12.5, leading=34 if is_zh else 24,
            alignment=TA_JUSTIFY, firstLineIndent=0,
            textColor=colors.HexColor("#151515"), wordWrap="CJK" if is_zh else None,
        )
    detail_paragraph_style = ParagraphStyle(
        "CertificateDetailBody", fontName=font_name if is_zh else english_font_name,
        fontSize=(15.5 if is_ntnu_tutor else 14) if is_zh else 12,
        leading=(34 if is_ntnu_tutor else 31) if is_zh else 22,
        alignment=TA_JUSTIFY, firstLineIndent=0,
        textColor=colors.HexColor("#151515"),
        wordWrap="CJK" if is_zh else None,
    )
    summary_lead_style = ParagraphStyle(
        "CertificateSummaryLead", fontName=font_name if is_zh else english_font_name,
        fontSize=(21 if is_ntnu_tutor else 18) if is_zh else (15 if is_ntnu_tutor else 13),
        leading=(34 if is_ntnu_tutor else 29) if is_zh else (20 if is_ntnu_tutor else 18),
        alignment=0, textColor=colors.HexColor("#151515"), wordWrap="CJK" if is_zh else None,
    )
    detail_lead_style = ParagraphStyle(
        "CertificateDetailLead", fontName=font_name if is_zh else english_font_name,
        fontSize=17 if is_zh else 13, leading=29 if is_zh else 18,
        alignment=0, textColor=colors.HexColor("#151515"), wordWrap="CJK" if is_zh else None,
    )
    small_style = ParagraphStyle(
        "CertificateCell", fontName=font_name, fontSize=9.5, leading=12,
        alignment=TA_CENTER, textColor=colors.HexColor("#221D19"),
    )

    ordered_fields = [key for key in ("date", "nationality", "level", "hours") if key in detail_fields]
    if is_zh:
        field_config = {
            "date": (f"日期<br/><font name='{english_font_name}' size='8'>Date</font>", 1.2),
            "nationality": (f"學生國籍<br/><font name='{english_font_name}' size='8'>Nationality</font>", 1.35),
            "level": (f"學生程度<br/><font name='{english_font_name}' size='8'>Chinese level</font>", 1.3),
            "hours": (f"時數<br/><font name='{english_font_name}' size='8'>Hours</font>", .75),
        }
    else:
        field_config = {
            "date": (f"<font name='{english_font_name}'>Date</font>", 1.2),
            "nationality": (f"<font name='{english_font_name}'>Nationality</font>", 1.35),
            "level": (f"<font name='{english_font_name}'>Chinese level</font>", 1.3),
            "hours": (f"<font name='{english_font_name}'>Hours</font>", .75),
        }
    detail_rows = []
    for section in data["sections"]:
        for row in section["rows"]:
            session = row["session"]
            values = {
                "date": session.class_date.strftime("%Y/%m/%d"),
                "nationality": row["student_nationality"],
                "level": row["student_level"],
                "hours": hours_text(session.duration),
            }
            detail_rows.append([values[key] for key in ordered_fields])

    # The revised template reserves the lower area for the formal issue date,
    # department logo, and seal. Six detail rows keep the table clear of the
    # enlarged, raised seal on every page, including continuation pages.
    rows_per_page = 6
    chunks = [detail_rows[index:index + rows_per_page] for index in range(0, len(detail_rows), rows_per_page)]
    if not chunks:
        chunks = [[]]
    if version == "summary":
        chunks = [None]

    overlay_buffer = BytesIO()
    pdf_canvas = canvas.Canvas(overlay_buffer, pagesize=A4)
    page_width, _ = A4
    total_pages = len(chunks)
    for page_index, chunk in enumerate(chunks, start=1):
        if is_zh:
            pdf_canvas.setFillColor(colors.HexColor("#151515"))
            pdf_canvas.setFont(bold_font_name, 22)
            pdf_canvas.drawCentredString(page_width / 2, 650, title)
        else:
            # A single English title stands alone (no Chinese title above it), so it gets a
            # larger size than the old secondary-line English title did.
            pdf_canvas.setFillColor(colors.HexColor("#151515"))
            pdf_canvas.setFont(english_bold_font_name, 18)
            pdf_canvas.drawCentredString(page_width / 2, 650, title)

        is_summary = version == "summary"
        paragraph_style = summary_paragraph_style if is_summary else detail_paragraph_style
        # Summary certificates read more formally with a narrower text block and
        # wider left/right margins. Detailed certificates keep the wider block so
        # the four-column table remains legible.
        paragraph_width = 425 if is_summary else 465
        paragraph_x = (page_width - paragraph_width) / 2
        paragraph_top = 600 if is_summary else 620
        if certificate_lead:
            lead_style = summary_lead_style if is_summary else detail_lead_style
            lead = Paragraph(certificate_lead, lead_style)
            # Generic certificate names may include two names plus a longer partner-program
            # ID. Give only this lead line the detailed-page width so the intentionally
            # grouped name/ID line does not wrap; the summary body keeps its wider margins.
            lead_width = 465 if is_summary and not is_ntnu_tutor else paragraph_width
            lead_x = (page_width - lead_width) / 2
            _, lead_height = lead.wrap(lead_width, 40)
            lead_y = paragraph_top - lead_height
            lead.drawOn(pdf_canvas, lead_x, lead_y)
            paragraph_top = lead_y - (16 if is_summary else 12)
        paragraph = Paragraph(certificate_paragraph, paragraph_style)
        _, paragraph_height = paragraph.wrap(paragraph_width, 180)
        paragraph_y = paragraph_top - paragraph_height
        paragraph.drawOn(pdf_canvas, paragraph_x, paragraph_y)

        if version == "detailed":
            if is_zh:
                label = "輔導時數明細" if page_index == 1 else "輔導時數明細（續）"
            else:
                label = "Hours Detail" if page_index == 1 else "Hours Detail (continued)"
            pdf_canvas.setFillColor(colors.HexColor("#392E26"))
            pdf_canvas.setFont(bold_font_name if is_zh else english_bold_font_name, 11.5)
            pdf_canvas.drawString(72, 450, label)
            page_number_style = ParagraphStyle(
                "CertificatePageNumber", fontName=bold_font_name, fontSize=11.5, leading=14,
                alignment=TA_RIGHT, textColor=colors.HexColor("#392E26"),
            )
            if is_zh:
                page_number_text = (
                    f"第 {page_index} 頁，共 {total_pages} 頁 / "
                    f'<font name="{english_bold_font_name}">Page {page_index} of {total_pages}</font>'
                )
            else:
                page_number_text = (
                    f'<font name="{english_bold_font_name}">Page {page_index} of {total_pages}</font>'
                )
            page_number = Paragraph(page_number_text, page_number_style)
            page_number.wrap(280, 16)
            page_number.drawOn(pdf_canvas, page_width - 352, 447)

            headers = [Paragraph(field_config[key][0], small_style) for key in ordered_fields]
            if chunk:
                body = [[Paragraph(mixed_font_markup(value), small_style) for value in row] for row in chunk]
            else:
                empty_text = "此期間沒有有效時數紀錄 / No verified records" if is_zh else "No verified records for this period"
                body = [[Paragraph(empty_text, small_style)] + [""] * (len(headers) - 1)]
            weights = [field_config[key][1] for key in ordered_fields]
            total_weight = sum(weights) or 1
            column_widths = [465 * weight / total_weight for weight in weights]
            table = Table([headers] + body, colWidths=column_widths, rowHeights=[34] + [29] * len(body))
            table_style = [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.91, 0.87, 0.82, alpha=.42)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#2A211B")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ]
            for row_index in range(1, len(body) + 1):
                if row_index % 2 == 0:
                    table_style.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.Color(0.96, 0.93, 0.89, alpha=.24)))
            if not chunk and len(headers) > 1:
                table_style.append(("SPAN", (0, 1), (-1, 1)))
            table.setStyle(TableStyle(table_style))
            _, table_height = table.wrap(465, 330)
            table.drawOn(pdf_canvas, 65, 433 - table_height)

        pdf_canvas.saveState()
        pdf_canvas.setFillColor(colors.HexColor("#171310"))
        generated_at = data["generated_at"]
        generated_date = (
            timezone.localtime(generated_at).date()
            if timezone.is_aware(generated_at)
            else generated_at.date()
        )
        date_text = roc_date(generated_date) if is_zh else gregorian_date(generated_date)
        date_font = bold_font_name if is_zh else english_bold_font_name
        date_font_size = 18 if is_zh else 15
        date_character_spacing = 4.2 if is_zh else 2.4
        date_width = pdf_canvas.stringWidth(date_text, date_font, date_font_size)
        date_width += date_character_spacing * max(len(date_text) - 1, 0)
        issue_date = pdf_canvas.beginText((page_width - date_width) / 2, 132)
        issue_date.setFont(date_font, date_font_size)
        issue_date.setCharSpace(date_character_spacing)
        if is_zh:
            issue_date.textLine(date_text)
        else:
            issue_date.textLine(date_text)
        pdf_canvas.drawText(issue_date)
        pdf_canvas.restoreState()

        stamp_path = settings.BASE_DIR / "assets/certificates/CSL stamp.png"
        if stamp_path.exists():
            stamp_width = 110
            stamp_height = stamp_width * 344 / 398
            pdf_canvas.drawImage(
                ImageReader(stamp_path),
                page_width - 70 - stamp_width,
                100,
                width=stamp_width,
                height=stamp_height,
                preserveAspectRatio=True,
                mask="auto",
            )
        pdf_canvas.showPage()
    pdf_canvas.save()

    overlay_buffer.seek(0)
    overlay_reader = PdfReader(overlay_buffer)
    writer = PdfWriter()
    for overlay_page in overlay_reader.pages:
        page = PdfReader(str(template_path)).pages[0]
        page.merge_page(overlay_page)
        writer.add_page(page)
    output = BytesIO()
    writer.write(output)
    return _restrict_copy_and_selection(output.getvalue())


EXPORT_COLUMNS = (
    ("student_id", "學號 Student ID", 13),
    ("name_zh", "中文姓名", 13),
    ("name_en", "英文姓名", 16),
    ("role", "身分 Role", 14),
    ("semester", "學期 Semester", 14),
    ("date", "日期 Date", 12),
    ("time", "時間 Time", 9),
    ("hours", "時數 Hours", 9),
    ("counterpart_id", "對方學號", 12),
    ("counterpart", "輔導對象", 18),
    ("status", "狀態 Status", 15),
)
EXPORT_FIELD_KEYS = tuple(key for key, _header, _width in EXPORT_COLUMNS)
EXPORT_HEADERS = [header for _key, header, _width in EXPORT_COLUMNS]

_SPREADSHEET_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _spreadsheet_safe_value(value):
    """Neutralize CSV/Excel formula injection (docs/VULNERABILITY_SCAN_IMPROVEMENTS.md
    batch 6 item 2): a cell that starts with =, +, -, or @ is interpreted as a formula by
    Excel/LibreOffice/Google Sheets when the export is opened, so a crafted student ID or
    name (e.g. "=HYPERLINK(...)") could run arbitrary formulas — including data
    exfiltration — on whoever opens the file. Prefixing a leading apostrophe is the
    standard escape both applications already understand as "force literal text"; it also
    stops openpyxl's own Cell.value setter from auto-detecting the string as a formula
    (which it does purely based on a leading "="), so this one prefix protects CSV and
    XLSX identically.
    """
    text = str(value)
    if text.startswith(_SPREADSHEET_FORMULA_PREFIXES):
        return "'" + text
    return text


def _spreadsheet_safe_rows(rows):
    return [[_spreadsheet_safe_value(value) for value in row] for row in rows]


def normalize_export_fields(fields=None):
    """Return valid export field keys in the system-defined column order."""
    if fields is None:
        return list(EXPORT_FIELD_KEYS)
    requested = set(fields)
    return [key for key in EXPORT_FIELD_KEYS if key in requested]


def _export_schema(fields=None):
    selected = normalize_export_fields(fields)
    return [(key, header, width) for key, header, width in EXPORT_COLUMNS if key in selected]


def _export_rows(users, *, starts_on=None, ends_on=None, fields=None, program=None):
    selected_fields = normalize_export_fields(fields)
    rows = []
    for user in users:
        participant = Q(pairing__tutor=user) if user.role == Role.TUTOR else Q(pairing__tutee=user)
        sessions = ClassSession.objects.filter(participant).select_related(
            "pairing__semester", "pairing__tutor", "pairing__tutee",
            "pairing__tutee__roster_entry__program",
        ).prefetch_related("attendances", "class_records", "confirmations", "makeup_review")
        if program is not None:
            sessions = sessions.filter(pairing__tutee__roster_entry__program=program)
        if starts_on:
            sessions = sessions.filter(class_date__gte=starts_on)
        if ends_on:
            sessions = sessions.filter(class_date__lte=ends_on)
        for session in sessions:
            counterpart = session.pairing.tutee if user.role == Role.TUTOR else session.pairing.tutor
            values = {
                "student_id": user.username,
                "name_zh": user.name_zh,
                "name_en": user.name_en,
                "role": user.get_role_display(),
                "semester": session.pairing.semester.name_zh,
                "date": str(session.class_date),
                "time": session.start_time.strftime("%H:%M"),
                "hours": str(session.duration),
                "counterpart_id": counterpart.username,
                "counterpart": counterpart.bilingual_name,
                "status": "有效 / Verified" if class_is_valid(session) else "未成立 / Incomplete",
            }
            rows.append([values[key] for key in selected_fields])
        if not sessions.exists():
            values = {key: "" for key in EXPORT_FIELD_KEYS}
            values.update({
                "student_id": user.username,
                "name_zh": user.name_zh,
                "name_en": user.name_en,
                "role": user.get_role_display(),
                "status": "尚無課程 / No classes",
            })
            rows.append([values[key] for key in selected_fields])
    return rows


def build_excel_xlsx(users, *, starts_on=None, ends_on=None, fields=None, program=None):
    """Create a real .xlsx workbook (openpyxl) with the same columns as build_excel_xml()."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    schema = _export_schema(fields)
    headers = [header for _key, header, _width in schema]
    rows = _export_rows(
        users, starts_on=starts_on, ends_on=ends_on,
        fields=[key for key, _header, _width in schema], program=program,
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "輔導資料"
    worksheet.append(headers)
    for row in _spreadsheet_safe_rows(rows):
        worksheet.append(row)
    header_fill = PatternFill(start_color="0F4C75", end_color="0F4C75", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for header_cell in worksheet[1]:
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment
    for row_cells in worksheet.iter_rows(min_row=2):
        for data_cell in row_cells:
            data_cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, (_key, _header, width) in enumerate(schema, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_export_csv(users, *, starts_on=None, ends_on=None, fields=None, program=None):
    """Create a CSV export with the same columns as build_excel_xlsx().

    Written with a UTF-8 BOM so Excel on Windows opens the Chinese headers/content correctly.
    """
    import codecs
    import csv as csv_module

    schema = _export_schema(fields)
    headers = [header for _key, header, _width in schema]
    rows = _export_rows(
        users, starts_on=starts_on, ends_on=ends_on,
        fields=[key for key, _header, _width in schema], program=program,
    )
    buffer = StringIO()
    writer = csv_module.writer(buffer)
    writer.writerow(headers)
    writer.writerows(_spreadsheet_safe_rows(rows))
    return codecs.BOM_UTF8 + buffer.getvalue().encode("utf-8")


def build_export_pdf(users, *, starts_on=None, ends_on=None, fields=None, program=None):
    """Administrative report with the same columns as build_excel_xlsx()/build_export_csv().

    This is a plain data table for internal review, not a personal certificate — it reuses the
    certificate module's registered CJK/Latin fonts but has none of the certificate's formal
    layout. reportlab's platypus SimpleDocTemplate handles pagination automatically since the
    row count is unbounded (unlike the fixed-length certificate detail table).
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    font_name, _bold_font_name, _english_font_name, _english_bold_font_name = _register_certificate_fonts()

    schema = _export_schema(fields)
    headers = [header for _key, header, _width in schema]
    rows = _export_rows(
        users, starts_on=starts_on, ends_on=ends_on,
        fields=[key for key, _header, _width in schema], program=program,
    )
    header_style = ParagraphStyle("ExportHeader", fontName=font_name, fontSize=8.5, leading=11, textColor=colors.white)
    cell_style = ParagraphStyle("ExportCell", fontName=font_name, fontSize=8, leading=10)
    title_style = ParagraphStyle("ExportTitle", fontName=font_name, fontSize=14, leading=18)

    header_row = [Paragraph(header, header_style) for header in headers]
    if rows:
        body_rows = [[Paragraph(str(value), cell_style) for value in row] for row in rows]
    else:
        body_rows = [
            [Paragraph("沒有符合條件的資料 / No matching data", cell_style)] + [""] * (len(headers) - 1)
        ]
    table = Table([header_row] + body_rows, repeatRows=1)
    table_style = [
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F4C75")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_index in range(1, len(body_rows) + 1):
        if row_index % 2 == 0:
            table_style.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F5F5F5")))
    table.setStyle(TableStyle(table_style))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=15 * mm, bottomMargin=15 * mm, leftMargin=10 * mm, rightMargin=10 * mm,
    )
    generated_at = timezone.localtime().strftime("%Y-%m-%d %H:%M")
    doc.build([
        Paragraph("華語實習暨輔導系統 資料匯出 / MPTS Data Export", title_style),
        Paragraph(f"匯出時間 / Generated at: {generated_at}", cell_style),
        Spacer(1, 8),
        table,
    ])
    return _restrict_copy_and_selection(buffer.getvalue())
