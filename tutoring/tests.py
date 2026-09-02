from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import openpyxl

from django.contrib import admin as django_admin
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, TestCase
from django.urls import reverse
from django.utils import timezone

from accounts.models import AuditLog, EducationLevel, IdentityCategory, PartnerProgram, Role, RosterEntry, User
from .forms import ClassRecordForm, HoursDownloadForm, ScheduleClassForm, SemesterCreateForm
from .reporting import build_excel_xlsx, build_export_csv, build_hours_pdf, tutor_available_programs, user_has_hour_records

from .models import (
    InvitationStatus,
    MatchingInvitation,
    MakeupReview,
    Pairing,
    PairingReleaseReason,
    PairingReleaseRequest,
    PairingReleaseStatus,
    PairingStatus,
    PairingMessage,
    QualificationDocument,
    QualificationStatus,
    Semester,
    TuteeProfile,
    TutorProfile,
    Attendance,
    ClassDocument,
    ClassRecord,
    ClassConfirmation,
    ClassAlert,
    ClassAlertReason,
    ClassAlertStatus,
    ClassSession,
    ConfirmationStatus,
    HourAdjustment,
    IncidentReport,
    IncidentReportCategory,
    IncidentReportStatus,
    MakeupReviewStatus,
    validate_class_document_file,
    validate_class_record_attachment,
    validate_qualification_file,
)
from .admin import (
    AttendanceAdmin,
    ClassConfirmationAdmin,
    ClassRecordAdmin,
    ClassSessionAdmin,
    MakeupReviewAdmin,
    MatchingInvitationAdmin,
    PairingAdmin,
    PairingReleaseRequestAdmin,
)
from .services import (
    active_semester,
    anonymous_tutee_candidates,
    anonymous_tutor_candidates,
    archive_expired_semesters,
    check_in,
    cancel_class_alert,
    cancel_invitation,
    class_is_valid,
    confirm_counterpart,
    create_admin_pairing,
    respond_to_invitation,
    resolve_class_alert,
    resolve_incident_report,
    review_makeup,
    report_class_alert,
    process_pending_pairing_releases,
    review_pairing_release_request,
    reschedule_class,
    schedule_classes,
    send_invitation,
    submit_incident_report,
    submit_pairing_release_request,
    submit_class_record,
    synchronize_matching_state,
    user_program,
    visible_class_document_programs,
    visible_class_documents,
)


def minimal_pdf_bytes():
    """A genuinely parseable single-page PDF (batch 6 item 1 added real content
    validation via pypdf, so a plain b"%PDF-1.4..." byte string with no actual PDF
    structure is no longer accepted wherever a validator actually runs, e.g. through a
    ModelForm's is_valid())."""
    from pypdf import PdfWriter

    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    buffer = BytesIO()
    writer.write(buffer)
    return buffer.getvalue()


class SemesterTests(TestCase):
    def test_end_date_must_follow_start_date(self):
        semester = Semester(
            name_zh="測試學期",
            name_en="Test semester",
            starts_on=date(2026, 9, 1),
            ends_on=date(2026, 8, 31),
        )
        with self.assertRaises(ValidationError):
            semester.full_clean()

    def test_create_form_asks_for_dates_and_program(self):
        self.assertEqual(
            list(SemesterCreateForm().fields),
            ["name_zh", "name_en", "starts_on", "ends_on", "program"],
        )
        self.assertTrue(SemesterCreateForm().fields["program"].required)

    def test_semester_is_automatically_archived_after_six_months(self):
        today = date(2026, 7, 19)
        old = Semester.objects.create(
            name_zh="舊學期", name_en="Old semester",
            starts_on=date(2025, 9, 1), ends_on=date(2026, 1, 18), is_active=True,
        )
        recent = Semester.objects.create(
            name_zh="半年內", name_en="Within six months",
            starts_on=date(2025, 7, 1), ends_on=date(2026, 1, 19), is_active=True,
        )
        self.assertEqual(archive_expired_semesters(today=today), 1)
        old.refresh_from_db()
        recent.refresh_from_db()
        self.assertFalse(old.is_active)
        self.assertTrue(recent.is_active)

    def test_admin_can_archive_an_ended_semester(self):
        today = timezone.localdate()
        semester = Semester.objects.create(
            name_zh="已結束學期", name_en="Ended semester",
            starts_on=today - timedelta(days=90), ends_on=today - timedelta(days=1),
            is_active=True,
        )
        admin = User.objects.create_superuser(username="SEMESTER-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(reverse("tutoring:archive_semester", args=[semester.pk]))
        self.assertRedirects(response, f"{reverse('accounts:dashboard')}#semesters")
        semester.refresh_from_db()
        self.assertFalse(semester.is_active)

    def test_admin_can_delete_semester_without_pairings(self):
        today = timezone.localdate()
        semester = Semester.objects.create(
            name_zh="打錯的學期", name_en="Mistaken semester",
            starts_on=today, ends_on=today + timedelta(days=60), is_active=True,
        )
        admin = User.objects.create_superuser(username="DELETE-SEM-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(reverse("tutoring:delete_semester", args=[semester.pk]))
        self.assertRedirects(response, f"{reverse('accounts:dashboard')}#semesters")
        self.assertFalse(Semester.objects.filter(pk=semester.pk).exists())
        self.assertTrue(AuditLog.objects.filter(event_type="SEMESTER_DELETED").exists())

    def test_admin_cannot_delete_semester_with_pairings(self):
        today = timezone.localdate()
        semester = Semester.objects.create(
            name_zh="已有配對的學期", name_en="Semester with pairings",
            starts_on=today, ends_on=today + timedelta(days=60), is_active=True,
        )
        ntnu = PartnerProgram.objects.get(code="NTNU")
        tutor_roster = RosterEntry.objects.create(
            student_id="DEL-SEM-TUTOR", name_zh="老師", role=Role.TUTOR,
            education_level=EducationLevel.MASTER, identity_category=IdentityCategory.LOCAL,
        )
        tutee_roster = RosterEntry.objects.create(
            student_id="DEL-SEM-TUTEE", name_zh="學生", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=ntnu,
        )
        tutor = User.objects.create_user(username="DEL-SEM-TUTOR", password="Password-2026", role=Role.TUTOR, roster_entry=tutor_roster)
        tutee = User.objects.create_user(username="DEL-SEM-TUTEE", password="Password-2026", role=Role.TUTEE, roster_entry=tutee_roster)
        Pairing.objects.create(semester=semester, tutor=tutor, tutee=tutee)

        admin = User.objects.create_superuser(username="DELETE-SEM-ADMIN2", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(reverse("tutoring:delete_semester", args=[semester.pk]))
        self.assertRedirects(response, f"{reverse('accounts:dashboard')}#semesters")
        self.assertTrue(Semester.objects.filter(pk=semester.pk).exists())

    def test_admin_can_edit_semester_dates(self):
        today = timezone.localdate()
        semester = Semester.objects.create(
            name_zh="待修正學期", name_en="Semester to fix",
            starts_on=today, ends_on=today + timedelta(days=60), is_active=True,
        )
        admin = User.objects.create_superuser(username="EDIT-SEM-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(
            reverse("tutoring:update_semester", args=[semester.pk]),
            {
                f"semester-{semester.pk}-name_zh": "修正後學期",
                f"semester-{semester.pk}-name_en": "Fixed semester",
                f"semester-{semester.pk}-starts_on": (today + timedelta(days=5)).isoformat(),
                f"semester-{semester.pk}-ends_on": (today + timedelta(days=70)).isoformat(),
                f"semester-{semester.pk}-is_active": "on",
            },
        )
        self.assertRedirects(response, f"{reverse('accounts:dashboard')}#semesters")
        semester.refresh_from_db()
        self.assertEqual(semester.name_zh, "修正後學期")
        self.assertEqual(semester.starts_on, today + timedelta(days=5))
        log = AuditLog.objects.get(event_type="SEMESTER_UPDATED")
        self.assertEqual(log.metadata["semester_id"], semester.pk)

    def test_dashboard_shows_edit_toggle_for_each_semester(self):
        today = timezone.localdate()
        semester = Semester.objects.create(
            name_zh="顯示編輯用學期", name_en="Semester for edit UI",
            starts_on=today, ends_on=today + timedelta(days=60), is_active=True,
        )
        admin = User.objects.create_superuser(username="EDIT-UI-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, reverse("tutoring:update_semester", args=[semester.pk]))
        self.assertContains(response, reverse("tutoring:delete_semester", args=[semester.pk]))


class ProgramScopedSemesterTests(TestCase):
    """MEETING_CHANGE_REQUIREMENTS_2026-08-04.md item 15: program-scoped, overlapping periods."""

    def setUp(self):
        self.ntnu = PartnerProgram.objects.get(code="NTNU")
        self.maryland = PartnerProgram.objects.get(code="MARYLAND")
        self.today = timezone.localdate()

    def test_more_than_three_active_future_semesters_can_be_created(self):
        for index in range(5):
            Semester.objects.create(
                name_zh=f"期間 {index}", name_en=f"Period {index}",
                starts_on=self.today + timedelta(days=400 * index),
                ends_on=self.today + timedelta(days=400 * index + 60),
                is_active=True, program=self.ntnu,
            ).full_clean()
        self.assertEqual(Semester.objects.filter(is_active=True).count(), 5)

    def test_same_program_overlap_is_rejected(self):
        Semester.objects.create(
            name_zh="NTNU 一", name_en="NTNU one", program=self.ntnu, is_active=True,
            starts_on=self.today, ends_on=self.today + timedelta(days=90),
        )
        overlapping = Semester(
            name_zh="NTNU 二", name_en="NTNU two", program=self.ntnu, is_active=True,
            starts_on=self.today + timedelta(days=30), ends_on=self.today + timedelta(days=120),
        )
        with self.assertRaises(ValidationError):
            overlapping.full_clean()

    def test_different_program_overlap_is_allowed(self):
        Semester.objects.create(
            name_zh="NTNU 學期", name_en="NTNU semester", program=self.ntnu, is_active=True,
            starts_on=self.today, ends_on=self.today + timedelta(days=90),
        )
        maryland_period = Semester(
            name_zh="馬里蘭計畫", name_en="Maryland program", program=self.maryland, is_active=True,
            starts_on=self.today, ends_on=self.today + timedelta(days=90),
        )
        maryland_period.full_clean()
        maryland_period.save()
        self.assertEqual(Semester.objects.filter(is_active=True, starts_on=self.today).count(), 2)

    def test_legacy_none_program_periods_still_reject_overlap_with_each_other(self):
        Semester.objects.create(
            name_zh="舊版一", name_en="Legacy one", is_active=True,
            starts_on=self.today, ends_on=self.today + timedelta(days=90),
        )
        overlapping = Semester(
            name_zh="舊版二", name_en="Legacy two", is_active=True,
            starts_on=self.today + timedelta(days=10), ends_on=self.today + timedelta(days=100),
        )
        with self.assertRaises(ValidationError):
            overlapping.full_clean()

    def test_active_semester_prefers_program_specific_then_falls_back_to_legacy(self):
        legacy = Semester.objects.create(
            name_zh="共用期間", name_en="Shared period", is_active=True,
            starts_on=self.today, ends_on=self.today + timedelta(days=90),
        )
        self.assertEqual(active_semester(program=self.maryland), legacy)
        ntnu_specific = Semester.objects.create(
            name_zh="NTNU 專屬", name_en="NTNU specific", program=self.ntnu, is_active=True,
            starts_on=self.today, ends_on=self.today + timedelta(days=90),
        )
        self.assertEqual(active_semester(program=self.ntnu), ntnu_specific)
        self.assertEqual(active_semester(program=self.maryland), legacy)
        self.assertEqual(active_semester(), legacy)

    def test_user_program_resolves_ordinary_tutor_to_ntnu_not_none(self):
        """An ordinary tutor (no explicit roster program) implicitly serves NTNU
        (tutor_can_serve_program()) — user_program() must resolve that to the real NTNU
        PartnerProgram, not bare None, or active_semester(program=user_program(tutor))
        silently falls back to looking up the legacy shared period instead of an
        NTNU-scoped one an Admin actually created (the bug this test guards against)."""
        tutee_roster = RosterEntry.objects.create(
            student_id="MULTI-TUTEE", name_zh="多計畫學生", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.ntnu,
        )
        tutee = User.objects.create_user(username="MULTI-TUTEE", password="Password-2026", role=Role.TUTEE, roster_entry=tutee_roster)
        tutor_roster = RosterEntry.objects.create(
            student_id="MULTI-TUTOR", name_zh="多計畫老師", role=Role.TUTOR,
            education_level=EducationLevel.MASTER, identity_category=IdentityCategory.LOCAL,
        )
        tutor = User.objects.create_user(username="MULTI-TUTOR", password="Password-2026", role=Role.TUTOR, roster_entry=tutor_roster)
        self.assertEqual(user_program(tutee), self.ntnu)
        self.assertEqual(user_program(tutor), self.ntnu)

    def test_ordinary_ntnu_tutor_sees_the_same_ntnu_specific_semester_as_ntnu_tutees(self):
        """Reproduces the real bug: creating an NTNU-scoped semester used to apply to NTNU
        tutees (whose roster program is explicitly NTNU) but not to ordinary NTNU tutors
        (whose roster program is None), because active_semester(program=None) only ever
        looks up the legacy shared period, never a program-specific one."""
        tutee_roster = RosterEntry.objects.create(
            student_id="SEM-BUG-TUTEE", name_zh="學期學生", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.ntnu,
        )
        tutee = User.objects.create_user(username="SEM-BUG-TUTEE", password="Password-2026", role=Role.TUTEE, roster_entry=tutee_roster)
        tutor_roster = RosterEntry.objects.create(
            student_id="SEM-BUG-TUTOR", name_zh="學期老師", role=Role.TUTOR,
            education_level=EducationLevel.MASTER, identity_category=IdentityCategory.LOCAL,
        )
        tutor = User.objects.create_user(username="SEM-BUG-TUTOR", password="Password-2026", role=Role.TUTOR, roster_entry=tutor_roster)

        ntnu_specific = Semester.objects.create(
            name_zh="NTNU 專屬學期", name_en="NTNU-specific semester", program=self.ntnu, is_active=True,
            starts_on=self.today, ends_on=self.today + timedelta(days=90),
        )
        self.assertEqual(active_semester(program=user_program(tutee)), ntnu_specific)
        self.assertEqual(active_semester(program=user_program(tutor)), ntnu_specific)

    def test_send_invitation_uses_tutee_program_period_over_unrelated_legacy_period(self):
        Semester.objects.filter(is_active=True).delete()
        legacy = Semester.objects.create(
            name_zh="共用期間", name_en="Shared period", is_active=True,
            starts_on=self.today - timedelta(days=400), ends_on=self.today - timedelta(days=300),
        )
        ntnu_period = Semester.objects.create(
            name_zh="NTNU 專屬", name_en="NTNU specific", program=self.ntnu, is_active=True,
            starts_on=self.today, ends_on=self.today + timedelta(days=90),
        )
        tutor_roster = RosterEntry.objects.create(
            student_id="INV-TUTOR", name_zh="老師", role=Role.TUTOR,
            education_level=EducationLevel.MASTER, identity_category=IdentityCategory.LOCAL,
        )
        tutor = User.objects.create_user(username="INV-TUTOR", password="Password-2026", role=Role.TUTOR, roster_entry=tutor_roster)
        TutorProfile.objects.create(tutor=tutor)
        QualificationDocument.objects.create(tutor=tutor, file="x.pdf", original_filename="x.pdf", status=QualificationStatus.APPROVED)
        tutee_roster = RosterEntry.objects.create(
            student_id="INV-TUTEE", name_zh="學生", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.ntnu,
        )
        tutee = User.objects.create_user(username="INV-TUTEE", password="Password-2026", role=Role.TUTEE, roster_entry=tutee_roster)
        invitation = send_invitation(initiator=tutor, tutor_id=tutor.pk, tutee_id=tutee.pk)
        self.assertEqual(invitation.semester_id, ntnu_period.pk)
        self.assertNotEqual(invitation.semester_id, legacy.pk)


class MatchingFixtureTestCase(TestCase):
    """Shared tutor/tutee fixtures and factory helpers. No test_ methods of its own — it exists
    so MatchingTests and MarylandTutorRosterTests can both use the same setUp() without either
    inheriting (and re-running) the other's tests."""

    def setUp(self):
        today = timezone.localdate()
        self.semester = Semester.objects.create(
            name_zh="115 學年度第 1 學期",
            name_en="Fall 2026",
            starts_on=today - timedelta(days=7),
            ends_on=today + timedelta(days=90),
            is_active=True,
        )
        self.ntnu_program = PartnerProgram.objects.get(code="NTNU")
        self.maryland_program = PartnerProgram.objects.get(code="MARYLAND")
        self.tutor = self.make_tutor("TUTOR100", "知名小老師", "Known Tutor")
        self.tutee = self.make_tutee("TUTEE100", "知名外籍生", "Known Tutee", self.ntnu_program)
        self.maryland = self.make_tutee("MARY100", "馬里蘭學生", "Maryland Student", self.maryland_program)
        self.maryland_tutor = self.make_maryland_tutor("MARYTUTOR100", "馬里蘭老師", "Maryland Tutor")

    def make_tutor(self, student_id, name_zh, name_en, program=None, education_level=EducationLevel.MASTER):
        roster = RosterEntry.objects.create(
            student_id=student_id,
            name_zh=name_zh,
            name_en=name_en,
            role=Role.TUTOR,
            education_level=education_level,
            identity_category=IdentityCategory.LOCAL,
            program=program,
        )
        user = User.objects.create_user(
            username=student_id, password="Matching-password-2026", role=Role.TUTOR,
            roster_entry=roster, name_zh=name_zh, name_en=name_en,
        )
        TutorProfile.objects.create(
            tutor=user, gender="MALE", native_language="Mandarin Chinese", nationality="Taiwan",
            level_listening=4, level_speaking=5, level_reading=4, level_writing=4,
            available_days=["MON"], available_time_slots=["13:00-15:00"],
        )
        QualificationDocument.objects.create(
            tutor=user,
            file=SimpleUploadedFile(f"{student_id}.pdf", b"%PDF-1.4 test"),
            original_filename=f"{student_id}.pdf",
            status=QualificationStatus.APPROVED,
        )
        return user

    def make_maryland_tutor(self, student_id, name_zh, name_en):
        """A tutor on the Maryland course roster (item 4): bachelor's level, program=MARYLAND."""
        return self.make_tutor(
            student_id, name_zh, name_en, program=self.maryland_program, education_level=EducationLevel.BACHELOR,
        )

    def make_tutee(self, student_id, name_zh, name_en, program):
        roster = RosterEntry.objects.create(
            student_id=student_id,
            name_zh=name_zh,
            name_en=name_en,
            role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE,
            identity_category=IdentityCategory.INTERNATIONAL,
            program=program,
        )
        user = User.objects.create_user(
            username=student_id, password="Matching-password-2026", role=Role.TUTEE,
            roster_entry=roster, name_zh=name_zh, name_en=name_en,
        )
        TuteeProfile.objects.create(
            tutee=user, gender="FEMALE", native_language="English", nationality="United States",
            department="Languages", overall_level="B1", learning_duration="1_TO_2_YEARS",
            target_skills=["SPEAKING"], skills_to_improve="希望加強日常會話",
            preferred_days=["TUE"], preferred_time_slots=["15:00-17:00"],
        )
        return user


class MatchingTests(MatchingFixtureTestCase):
    def test_find_students_uses_backend_program_name(self):
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, self.ntnu_program.name_zh)
        self.assertContains(response, self.ntnu_program.name_en)

        self.client.force_login(self.maryland_tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, self.maryland_program.name_zh)
        self.assertContains(response, self.maryland_program.name_en)

    def test_tutor_qualification_panel_is_full_width_and_reuses_upload_guidance(self):
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, 'class="panel status-panel"', html=False)
        self.assertNotContains(response, 'class="panel status-panel focused-panel"', html=False)
        self.assertContains(response, "請擇一上傳以下文件")
        self.assertContains(response, "Please upload one of the following documents")

    def test_invite_tutee_returns_to_find_tutee_tab(self):
        """The invite button lives on the dashboard's #find-tutee tab; redirecting back
        to a bare #overview after sending an invite (the old behaviour) dropped the
        tutor off the candidate list they were just browsing."""
        self.client.force_login(self.tutor)
        response = self.client.post(reverse("tutoring:invite_tutee", args=[self.tutee.pk]))
        self.assertRedirects(response, reverse("accounts:dashboard") + "#find-tutee")

    def test_invite_tutor_returns_to_find_tutor_tab(self):
        self.client.force_login(self.maryland)
        response = self.client.post(reverse("tutoring:invite_tutor", args=[self.maryland_tutor.pk]))
        self.assertRedirects(response, reverse("accounts:dashboard") + "#find-tutor")

    def test_respond_invitation_returns_to_invitations_tab(self):
        invitation = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        self.client.force_login(self.tutee)
        response = self.client.post(
            reverse("tutoring:respond_invitation", args=[invitation.pk]), {"action": "reject"}
        )
        self.assertRedirects(response, reverse("accounts:dashboard") + "#invitations")

    def test_cancel_pending_invitation_returns_to_invitations_tab(self):
        invitation = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        self.client.force_login(self.tutor)
        response = self.client.post(reverse("tutoring:cancel_invitation", args=[invitation.pk]))
        self.assertRedirects(response, reverse("accounts:dashboard") + "#invitations")

    def test_send_invitation_writes_audit_log(self):
        invitation = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        log = AuditLog.objects.get(event_type="INVITATION_SENT")
        self.assertEqual(log.actor, self.tutor)
        self.assertEqual(log.target_user, self.tutee)
        self.assertEqual(log.metadata["invitation_id"], invitation.pk)

    def test_respond_invitation_accept_writes_audit_log(self):
        invitation = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        pairing = respond_to_invitation(invitation_id=invitation.pk, responder=self.tutee, accept=True)
        log = AuditLog.objects.get(event_type="INVITATION_ACCEPTED")
        self.assertEqual(log.actor, self.tutee)
        self.assertEqual(log.target_user, self.tutee)
        self.assertEqual(log.metadata["pairing_id"], pairing.pk)

    def test_respond_invitation_reject_writes_audit_log(self):
        invitation = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        respond_to_invitation(invitation_id=invitation.pk, responder=self.tutee, accept=False)
        log = AuditLog.objects.get(event_type="INVITATION_REJECTED")
        self.assertEqual(log.actor, self.tutee)
        self.assertEqual(log.metadata["invitation_id"], invitation.pk)

    def test_cancel_invitation_writes_audit_log(self):
        invitation = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        cancel_invitation(invitation_id=invitation.pk, actor=self.tutor)
        log = AuditLog.objects.get(event_type="INVITATION_CANCELLED")
        self.assertEqual(log.actor, self.tutor)
        self.assertEqual(log.metadata["invitation_id"], invitation.pk)

    def test_expired_invitation_writes_audit_log_with_no_actor(self):
        invitation = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        invitation.expires_at = timezone.now() - timedelta(minutes=1)
        invitation.save(update_fields=["expires_at"])
        synchronize_matching_state()
        invitation.refresh_from_db()
        self.assertEqual(invitation.status, "EXPIRED")
        log = AuditLog.objects.get(event_type="INVITATION_EXPIRED")
        self.assertIsNone(log.actor)
        self.assertEqual(log.target_user, self.tutee)
        self.assertEqual(log.metadata["invitation_id"], invitation.pk)

    def test_accepting_invitation_auto_cancels_tutees_other_invitation_with_audit_log(self):
        other_tutor = self.make_tutor("MULTI-INV-TUTOR", "另一位老師", "Other Tutor")
        invitation_a = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        invitation_b = send_invitation(initiator=other_tutor, tutor_id=other_tutor.pk, tutee_id=self.tutee.pk)
        respond_to_invitation(invitation_id=invitation_a.pk, responder=self.tutee, accept=True)
        invitation_b.refresh_from_db()
        self.assertEqual(invitation_b.status, InvitationStatus.CANCELLED)
        log = AuditLog.objects.get(event_type="INVITATION_AUTO_CANCELLED")
        self.assertIsNone(log.actor)
        self.assertEqual(log.target_user, self.tutee)
        self.assertEqual(log.metadata["invitation_id"], invitation_b.pk)

    def test_dashboard_shows_resolved_invitations_in_history_not_pending_lists(self):
        invitation = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        cancel_invitation(invitation_id=invitation.pk, actor=self.tutor)
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertEqual(response.context["sent_invitations"], [])
        history_statuses = [row["status"] for row in response.context["invitation_history"]]
        self.assertIn("CANCELLED", history_statuses)
        self.assertContains(response, "已取消 / Cancelled")

    def test_ntnu_tutee_dashboard_hides_permanently_empty_sent_invitation_card(self):
        """NTNU tutees can never initiate invitations (allow_tutee_initiate_invitation=False),
        so their "已發送的邀請 / Sent" card is permanently empty dead UI and should be hidden
        entirely, unlike Maryland tutees who can send invitations."""
        self.client.force_login(self.tutee)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertFalse(response.context["is_maryland"])
        self.assertNotContains(response, "已發送的邀請")

    def test_maryland_tutee_dashboard_shows_sent_invitation_card(self):
        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertTrue(response.context["is_maryland"])
        self.assertContains(response, "已發送的邀請")

    def test_candidate_cards_flag_test_prefixed_accounts(self):
        """TEST- prefixed student IDs (the project's established convention for QA
        fixtures, see docs/SECURITY_CHECKLIST.md's note on seed_test_roster.py) get an
        is_test hint on their anonymous candidate card so testers can tell them apart
        from real students — without exposing the actual student ID itself."""
        test_tutee = self.make_tutee("TEST-CANDIDATE1", "測試學生", "Test Tutee", self.ntnu_program)
        candidates = anonymous_tutee_candidates(semester=self.semester, tutor=self.tutor)
        by_id = {c["user_id"]: c for c in candidates}
        self.assertTrue(by_id[test_tutee.pk]["is_test"])
        self.assertFalse(by_id[self.tutee.pk]["is_test"])

    def test_tutee_find_teacher_heading_stays_csl_teacher(self):
        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "華語老師")
        self.assertContains(response, "Chinese teachers")

    def test_anonymous_candidate_data_excludes_identity(self):
        candidate = anonymous_tutee_candidates(semester=self.semester, tutor=self.tutor)[0]
        self.assertNotIn("name_zh", candidate)
        self.assertNotIn("name_en", candidate)
        self.assertNotIn("student_id", candidate)
        self.assertNotIn("phone", candidate)
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertNotContains(response, "知名外籍生")
        self.assertNotContains(response, "Known Tutee")
        self.assertNotContains(response, "TUTEE100")
        self.assertContains(response, "United States")
        self.assertContains(response, "TOCFL B1")

    def test_tutee_candidates_can_be_filtered_by_compound_criteria(self):
        TuteeProfile.objects.filter(tutee=self.maryland).delete()
        other = self.make_tutee("TUTEE200", "另一位外籍生", "Other Tutee", self.ntnu_program)
        other.tutee_profile.gender = "MALE"
        other.tutee_profile.native_language = "Korean"
        other.tutee_profile.overall_level = "A1"
        other.tutee_profile.target_skills = ["LISTENING", "READING"]
        other.tutee_profile.preferred_days = ["WED"]
        other.tutee_profile.preferred_time_slots = ["09:00-11:00"]
        other.tutee_profile.save()

        all_ids = {c["user_id"] for c in anonymous_tutee_candidates(semester=self.semester, tutor=self.tutor)}
        self.assertEqual(all_ids, {self.tutee.pk, other.pk})

        gender_filtered = anonymous_tutee_candidates(
            semester=self.semester, tutor=self.tutor, filters={"gender": "MALE"}
        )
        self.assertEqual({c["user_id"] for c in gender_filtered}, {other.pk})

        level_filtered = anonymous_tutee_candidates(
            semester=self.semester, tutor=self.tutor, filters={"overall_level": "B1"}
        )
        self.assertEqual({c["user_id"] for c in level_filtered}, {self.tutee.pk})

        language_filtered = anonymous_tutee_candidates(
            semester=self.semester, tutor=self.tutor, filters={"native_language": "Korean"}
        )
        self.assertEqual({c["user_id"] for c in language_filtered}, {other.pk})

        skills_filtered = anonymous_tutee_candidates(
            semester=self.semester, tutor=self.tutor, filters={"target_skills": ["LISTENING", "READING"]}
        )
        self.assertEqual({c["user_id"] for c in skills_filtered}, {other.pk})

        days_filtered = anonymous_tutee_candidates(
            semester=self.semester, tutor=self.tutor, filters={"days": ["TUE"]}
        )
        self.assertEqual({c["user_id"] for c in days_filtered}, {self.tutee.pk})

        slots_filtered = anonymous_tutee_candidates(
            semester=self.semester, tutor=self.tutor, filters={"time_slots": ["09:00-11:00"]}
        )
        self.assertEqual({c["user_id"] for c in slots_filtered}, {other.pk})

        combined_filtered = anonymous_tutee_candidates(
            semester=self.semester,
            tutor=self.tutor,
            filters={"gender": "MALE", "native_language": "korean", "days": ["MON"]},
        )
        self.assertEqual(combined_filtered, [])

        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"), {"tutee_gender": "MALE"})
        self.assertContains(response, "1 位符合")
        self.assertContains(response, reverse("tutoring:invite_tutee", args=[other.pk]))
        self.assertNotContains(response, reverse("tutoring:invite_tutee", args=[self.tutee.pk]))

    def test_dashboard_greeting_falls_back_to_english_name_not_student_id(self):
        """International students often have no Chinese name (RosterEntry.name_zh /
        User.name_zh can be blank). The greeting and sidebar identity used to fall back
        straight to request.user.username (the student ID) whenever name_zh was blank,
        skipping name_en entirely — User.bilingual_name already has the right fallback
        chain (name_zh, else name_en, else username) and should be used instead."""
        no_chinese_name_tutee = self.make_tutee("NO-ZH-NAME", "", "English Only Name", self.ntnu_program)
        self.client.force_login(no_chinese_name_tutee)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "你好，English Only Name！")
        self.assertNotContains(response, "你好，NO-ZH-NAME！")

    def test_maryland_dashboard_hides_tutor_identity(self):
        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "知名小老師")
        self.assertNotContains(response, "Known Tutor")
        self.assertNotContains(response, "TUTOR100")
        self.assertContains(response, "Mandarin Chinese")

    def test_tutor_candidates_can_be_filtered_by_compound_criteria(self):
        other_tutor = self.make_maryland_tutor("TUTOR200", "另一位老師", "Other Tutor")
        other_tutor.tutor_profile.gender = "FEMALE"
        other_tutor.tutor_profile.native_language = "Spanish"
        other_tutor.tutor_profile.available_days = ["WED"]
        other_tutor.tutor_profile.available_time_slots = ["09:00-11:00"]
        other_tutor.tutor_profile.save()

        all_ids = {c["user_id"] for c in anonymous_tutor_candidates(semester=self.semester, tutee=self.maryland)}
        self.assertEqual(all_ids, {self.maryland_tutor.pk, other_tutor.pk})

        gender_filtered = anonymous_tutor_candidates(
            semester=self.semester, tutee=self.maryland, filters={"gender": "FEMALE"}
        )
        self.assertEqual({c["user_id"] for c in gender_filtered}, {other_tutor.pk})

        language_filtered = anonymous_tutor_candidates(
            semester=self.semester, tutee=self.maryland, filters={"native_language": "Spanish"}
        )
        self.assertEqual({c["user_id"] for c in language_filtered}, {other_tutor.pk})

        days_filtered = anonymous_tutor_candidates(
            semester=self.semester, tutee=self.maryland, filters={"days": ["MON"]}
        )
        self.assertEqual({c["user_id"] for c in days_filtered}, {self.maryland_tutor.pk})

        slots_filtered = anonymous_tutor_candidates(
            semester=self.semester, tutee=self.maryland, filters={"time_slots": ["09:00-11:00"]}
        )
        self.assertEqual({c["user_id"] for c in slots_filtered}, {other_tutor.pk})

        combined_filtered = anonymous_tutor_candidates(
            semester=self.semester,
            tutee=self.maryland,
            filters={"gender": "FEMALE", "days": ["MON"]},
        )
        self.assertEqual(combined_filtered, [])

        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:dashboard"), {"tutor_gender": "FEMALE"})
        self.assertContains(response, "1 位符合")
        self.assertContains(response, reverse("tutoring:invite_tutor", args=[other_tutor.pk]))
        self.assertNotContains(response, reverse("tutoring:invite_tutor", args=[self.maryland_tutor.pk]))

    def test_pending_candidate_uses_short_status_label(self):
        send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "待回覆")
        self.assertNotContains(response, "已有待回覆邀請 / Pending")

    def test_tutee_can_expand_anonymous_teacher_information_from_received_invitation(self):
        self.tutor.tutor_profile.teaching_notes = "重視生活會話與發音練習"
        self.tutor.tutor_profile.save(update_fields=["teaching_notes", "updated_at"])
        self.tutor.email = "known.tutor@example.com"
        self.tutor.save(update_fields=["email"])
        send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        self.client.force_login(self.tutee)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "查看老師資料")
        self.assertContains(response, "聽力教學")
        self.assertContains(response, "13:00-15:00")
        self.assertContains(response, "重視生活會話與發音練習")
        self.assertNotContains(response, "知名小老師")
        self.assertNotContains(response, "Known Tutor")
        self.assertNotContains(response, "TUTOR100")
        # Email is only disclosed after pairing, not in the anonymous invitation view.
        self.assertNotContains(response, "known.tutor@example.com")

    def test_active_pair_can_open_each_others_full_profile(self):
        Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        self.tutee.email = "known.tutee@example.com"
        self.tutee.save(update_fields=["email"])
        self.tutor.email = "known.tutor@example.com"
        self.tutor.save(update_fields=["email"])

        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:matched_profile", args=[self.tutee.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "知名外籍生")
        self.assertContains(response, "學習資料")
        self.assertContains(response, "希望加強日常會話")
        # Email is shown once a pairing is active.
        self.assertContains(response, "known.tutee@example.com")

        self.client.force_login(self.tutee)
        response = self.client.get(reverse("accounts:matched_profile", args=[self.tutor.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "知名小老師")
        self.assertContains(response, "教學資料")
        self.assertContains(response, "known.tutor@example.com")

    def test_unmatched_user_cannot_open_matched_profile(self):
        Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:matched_profile", args=[self.tutor.pk]))
        self.assertEqual(response.status_code, 404)

    def test_current_match_card_links_to_profile_and_message_thread(self):
        pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        self.client.force_login(self.tutee)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, reverse("accounts:matched_profile", args=[self.tutor.pk]))
        self.assertContains(response, "私訊 / Message")
        self.assertContains(response, reverse("tutoring:pairing_messages", args=[pairing.pk]))
        self.assertNotContains(response, "電話 / Phone")

    def test_ntnu_tutee_hides_hours_but_maryland_keeps_them(self):
        Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        self.client.force_login(self.tutee)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertNotContains(response, "已排時數 / Reserved")
        self.assertNotContains(response, "有效時數 / Verified")

        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "已排時數 / Reserved")
        self.assertContains(response, "有效時數 / Verified")

    def test_admin_matching_summary_renders(self):
        admin = User.objects.create_superuser(username="MATCH-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Active matches")

    def test_ntnu_tutee_cannot_initiate_but_maryland_can(self):
        with self.assertRaises(ValidationError):
            send_invitation(initiator=self.tutee, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        invitation = send_invitation(
            initiator=self.maryland, tutor_id=self.maryland_tutor.pk, tutee_id=self.maryland.pk
        )
        self.assertEqual(invitation.status, InvitationStatus.PENDING)
        self.assertEqual(invitation.initiated_by, self.maryland)

    def test_acceptance_creates_pairing_and_tutee_cannot_have_two_tutors(self):
        invitation = send_invitation(
            initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk
        )
        pairing = respond_to_invitation(invitation_id=invitation.pk, responder=self.tutee, accept=True)
        self.assertEqual(pairing.status, PairingStatus.ACTIVE)
        other_tutor = self.make_tutor("TUTOR200", "第二位老師", "Second Tutor")
        with self.assertRaises(ValidationError):
            send_invitation(initiator=other_tutor, tutor_id=other_tutor.pk, tutee_id=self.tutee.pk)

    def test_tutor_capacity_is_two_active_tutees(self):
        second = self.make_tutee("TUTEE200", "第二位學生", "Second Tutee", self.ntnu_program)
        third = self.make_tutee("TUTEE300", "第三位學生", "Third Tutee", self.ntnu_program)
        Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=second)
        with self.assertRaises(ValidationError):
            send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=third.pk)

    def test_invitation_expires_in_five_days(self):
        invitation = send_invitation(
            initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk
        )
        remaining = invitation.expires_at - timezone.now()
        self.assertGreater(remaining, timedelta(days=4, hours=23))
        self.assertLessEqual(remaining, timedelta(days=5))

    def test_auto_eligible_release_ends_pairing_after_48_hours_and_cancels_future_class(self):
        pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        class_session = ClassSession.objects.create(
            pairing=pairing,
            class_date=timezone.localdate() + timedelta(days=7),
            start_time=time(15, 0),
            duration=1,
            created_by=self.tutor,
        )
        requested_at = timezone.now()
        release_request = submit_pairing_release_request(
            pairing_id=pairing.pk,
            requester=self.tutee,
            reason=PairingReleaseReason.SCHEDULE_CONFLICT,
            now=requested_at,
        )
        self.assertAlmostEqual(
            release_request.auto_resolve_at,
            requested_at + timedelta(hours=48),
            delta=timedelta(seconds=1),
        )
        self.assertEqual(
            process_pending_pairing_releases(now=requested_at + timedelta(hours=47, minutes=59)), 0
        )
        release_request.refresh_from_db()
        self.assertEqual(release_request.status, PairingReleaseStatus.PENDING)
        self.assertEqual(process_pending_pairing_releases(now=requested_at + timedelta(hours=48, seconds=1)), 1)
        pairing.refresh_from_db()
        release_request.refresh_from_db()
        class_session.refresh_from_db()
        self.assertEqual(pairing.status, PairingStatus.ENDED)
        self.assertEqual(release_request.status, PairingReleaseStatus.AUTO_APPROVED)
        self.assertEqual(class_session.status, "CANCELLED")
        replacement_tutor = self.make_tutor("TUTOR-REPLACEMENT", "新老師", "Replacement Tutor")
        replacement_invitation = send_invitation(
            initiator=replacement_tutor,
            tutor_id=replacement_tutor.pk,
            tutee_id=self.tutee.pk,
        )
        self.assertEqual(replacement_invitation.status, InvitationStatus.PENDING)
        with self.assertRaises(ValidationError):
            send_invitation(
                initiator=self.tutor,
                tutor_id=self.tutor.pk,
                tutee_id=self.tutee.pk,
            )

    def test_conduct_release_never_auto_approves_and_admin_can_approve(self):
        pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        requested_at = timezone.now()
        with self.assertRaises(ValidationError):
            submit_pairing_release_request(
                pairing_id=pairing.pk,
                requester=self.tutor,
                reason=PairingReleaseReason.CONDUCT,
                note="",
                now=requested_at,
            )
        release_request = submit_pairing_release_request(
            pairing_id=pairing.pk,
            requester=self.tutor,
            reason=PairingReleaseReason.CONDUCT,
            note="課堂互動有不適當行為",
            now=requested_at,
        )
        self.assertIsNone(release_request.auto_resolve_at)
        self.assertEqual(process_pending_pairing_releases(now=requested_at + timedelta(days=10)), 0)
        pairing.refresh_from_db()
        self.assertEqual(pairing.status, PairingStatus.ACTIVE)
        admin = User.objects.create_superuser(username="RELEASE-ADMIN", password="Admin-password-2026")
        review_pairing_release_request(
            request_id=release_request.pk,
            admin=admin,
            approve=True,
            note="已確認雙方狀況",
        )
        pairing.refresh_from_db()
        release_request.refresh_from_db()
        self.assertEqual(pairing.status, PairingStatus.ENDED)
        self.assertEqual(release_request.status, PairingReleaseStatus.APPROVED)

    def test_rejected_release_keeps_pairing_active(self):
        pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        release_request = submit_pairing_release_request(
            pairing_id=pairing.pk,
            requester=self.tutee,
            reason=PairingReleaseReason.UNREACHABLE,
        )
        admin = User.objects.create_superuser(username="REJECT-ADMIN", password="Admin-password-2026")
        review_pairing_release_request(
            request_id=release_request.pk,
            admin=admin,
            approve=False,
            note="已恢復聯繫",
        )
        pairing.refresh_from_db()
        release_request.refresh_from_db()
        self.assertEqual(pairing.status, PairingStatus.ACTIVE)
        self.assertEqual(release_request.status, PairingReleaseStatus.REJECTED)

    def test_release_request_appears_for_participant_and_admin(self):
        pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "申請解除配對")
        self.assertNotContains(response, "管理員 48 小時內未處理，系統將自動解除")
        response = self.client.post(
            reverse("tutoring:request_pairing_release", args=[pairing.pk]),
            {"reason": PairingReleaseReason.NO_SHOW, "note": "已多次未到"},
        )
        self.assertRedirects(response, reverse("accounts:dashboard") + "#overview")
        self.assertTrue(PairingReleaseRequest.objects.filter(pairing=pairing).exists())
        admin = User.objects.create_superuser(username="DASHBOARD-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "解除配對審核")
        self.assertContains(response, "已多次未到")

    def test_pending_invitation_cap_blocks_new_invitations_for_tutor(self):
        second = self.make_tutee("TUTEE210", "第二位學生", "Second Tutee", self.ntnu_program)
        third = self.make_tutee("TUTEE220", "第三位學生", "Third Tutee", self.ntnu_program)
        fourth = self.make_tutee("TUTEE230", "第四位學生", "Fourth Tutee", self.ntnu_program)
        send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=second.pk)
        send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=third.pk)
        with self.assertRaises(ValidationError):
            send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=fourth.pk)

    def test_pending_invitation_cap_counts_both_directions_for_tutee(self):
        tutor_b = self.make_maryland_tutor("TUTOR210", "第二位老師", "Second Tutor")
        tutor_c = self.make_maryland_tutor("TUTOR220", "第三位老師", "Third Tutor")
        tutor_d = self.make_maryland_tutor("TUTOR230", "第四位老師", "Fourth Tutor")
        send_invitation(initiator=self.maryland_tutor, tutor_id=self.maryland_tutor.pk, tutee_id=self.maryland.pk)
        send_invitation(initiator=tutor_b, tutor_id=tutor_b.pk, tutee_id=self.maryland.pk)
        send_invitation(initiator=self.maryland, tutor_id=tutor_c.pk, tutee_id=self.maryland.pk)
        with self.assertRaises(ValidationError):
            send_invitation(initiator=self.maryland, tutor_id=tutor_d.pk, tutee_id=self.maryland.pk)

    def test_tutor_reaching_capacity_cancels_other_pending_invitations(self):
        tutee_b = self.make_tutee("TUTEE240", "乙學生", "Tutee B", self.ntnu_program)
        tutee_c = self.make_tutee("TUTEE250", "丙學生", "Tutee C", self.ntnu_program)
        invitation_a = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        invitation_b = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=tutee_b.pk)
        invitation_c = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=tutee_c.pk)

        respond_to_invitation(invitation_id=invitation_a.pk, responder=self.tutee, accept=True)
        invitation_c.refresh_from_db()
        self.assertEqual(invitation_c.status, InvitationStatus.PENDING)

        respond_to_invitation(invitation_id=invitation_b.pk, responder=tutee_b, accept=True)
        invitation_c.refresh_from_db()
        self.assertEqual(invitation_c.status, InvitationStatus.CANCELLED)

    def test_maryland_initiated_invitation_cancels_tutees_other_pending_on_acceptance(self):
        tutor_b = self.make_maryland_tutor("TUTOR240", "乙老師", "Tutor B")
        invitation_to_tutor = send_invitation(
            initiator=self.maryland, tutor_id=self.maryland_tutor.pk, tutee_id=self.maryland.pk
        )
        invitation_to_tutor_b = send_invitation(
            initiator=self.maryland, tutor_id=tutor_b.pk, tutee_id=self.maryland.pk
        )
        respond_to_invitation(invitation_id=invitation_to_tutor.pk, responder=self.maryland_tutor, accept=True)
        invitation_to_tutor_b.refresh_from_db()
        self.assertEqual(invitation_to_tutor_b.status, InvitationStatus.CANCELLED)


class MarylandTutorRosterTests(MatchingFixtureTestCase):
    """MEETING_CHANGE_REQUIREMENTS_2026-08-04.md item 4: Maryland tutees only match tutors on
    the Maryland course roster, and only bachelor's-level ones at that; ordinary tutors keep
    serving NTNU tutees as before."""

    def test_ordinary_tutor_does_not_see_maryland_tutee_as_candidate(self):
        candidates = anonymous_tutee_candidates(semester=self.semester, tutor=self.tutor)
        self.assertNotIn(self.maryland.pk, {c["user_id"] for c in candidates})
        self.assertIn(self.tutee.pk, {c["user_id"] for c in candidates})

    def test_maryland_tutor_only_sees_maryland_tutee_not_ntnu(self):
        candidates = anonymous_tutee_candidates(semester=self.semester, tutor=self.maryland_tutor)
        ids = {c["user_id"] for c in candidates}
        self.assertIn(self.maryland.pk, ids)
        self.assertNotIn(self.tutee.pk, ids)

    def test_maryland_roster_tutor_at_wrong_education_level_is_excluded(self):
        masters_tutor = self.make_tutor(
            "MARY-MASTERS", "碩士老師", "Masters Tutor",
            program=self.maryland_program, education_level=EducationLevel.MASTER,
        )
        tutee_candidates = anonymous_tutor_candidates(semester=self.semester, tutee=self.maryland)
        self.assertNotIn(masters_tutor.pk, {c["user_id"] for c in tutee_candidates})
        with self.assertRaises(ValidationError):
            send_invitation(initiator=self.maryland, tutor_id=masters_tutor.pk, tutee_id=self.maryland.pk)

    def test_send_invitation_rejects_ordinary_tutor_for_maryland_tutee_even_off_screen(self):
        with self.assertRaises(ValidationError):
            send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.maryland.pk)

    def test_send_invitation_rejects_maryland_tutor_for_ntnu_tutee_even_off_screen(self):
        with self.assertRaises(ValidationError):
            send_invitation(initiator=self.maryland_tutor, tutor_id=self.maryland_tutor.pk, tutee_id=self.tutee.pk)

    def test_ordinary_tutor_can_still_pair_with_ntnu_tutee(self):
        invitation = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        self.assertEqual(invitation.status, InvitationStatus.PENDING)


class AdminPairingTests(MatchingFixtureTestCase):
    """MEETING_CHANGE_REQUIREMENTS_2026-08-04.md item 12: Admin can build a pairing directly,
    including one extra active tutee for non-NTNU programs — but never for NTNU."""

    def test_non_admin_cannot_create_admin_pairing(self):
        with self.assertRaises(ValidationError):
            create_admin_pairing(
                admin=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk, semester_id=self.semester.pk,
            )

    def test_admin_can_create_pairing_without_invitation(self):
        admin = User.objects.create_superuser(username="PAIR-ADMIN1", password="Admin-password-2026")
        pairing = create_admin_pairing(
            admin=admin, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk, semester_id=self.semester.pk,
        )
        self.assertEqual(pairing.status, PairingStatus.ACTIVE)
        self.assertEqual(pairing.created_by, admin)
        self.assertIsNone(pairing.invitation)
        log = AuditLog.objects.get(event_type="ADMIN_PAIRING_CREATED")
        self.assertEqual(log.metadata["tutor"], self.tutor.username)
        self.assertEqual(log.metadata["tutee"], self.tutee.username)

    def test_admin_can_grant_third_tutee_for_non_ntnu_program(self):
        admin = User.objects.create_superuser(username="PAIR-ADMIN2", password="Admin-password-2026")
        second_maryland_tutee = self.make_tutee("MARY200", "馬里蘭學生二", "Maryland Student 2", self.maryland_program)
        third_maryland_tutee = self.make_tutee("MARY300", "馬里蘭學生三", "Maryland Student 3", self.maryland_program)
        invitation_a = send_invitation(
            initiator=self.maryland_tutor, tutor_id=self.maryland_tutor.pk, tutee_id=self.maryland.pk,
        )
        invitation_b = send_invitation(
            initiator=self.maryland_tutor, tutor_id=self.maryland_tutor.pk, tutee_id=second_maryland_tutee.pk,
        )
        respond_to_invitation(invitation_id=invitation_a.pk, responder=self.maryland, accept=True)
        respond_to_invitation(invitation_id=invitation_b.pk, responder=second_maryland_tutee, accept=True)
        with self.assertRaises(ValidationError):
            send_invitation(
                initiator=self.maryland_tutor, tutor_id=self.maryland_tutor.pk, tutee_id=third_maryland_tutee.pk,
            )
        pairing = create_admin_pairing(
            admin=admin, tutor_id=self.maryland_tutor.pk, tutee_id=third_maryland_tutee.pk,
            semester_id=self.semester.pk,
        )
        self.assertEqual(pairing.status, PairingStatus.ACTIVE)
        self.assertEqual(
            Pairing.objects.filter(tutor=self.maryland_tutor, status=PairingStatus.ACTIVE).count(), 3
        )

    def test_admin_cannot_grant_third_tutee_for_ntnu(self):
        admin = User.objects.create_superuser(username="PAIR-ADMIN3", password="Admin-password-2026")
        tutee_b = self.make_tutee("TUTEE300", "乙學生", "Tutee B", self.ntnu_program)
        tutee_c = self.make_tutee("TUTEE310", "丙學生", "Tutee C", self.ntnu_program)
        invitation_a = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk)
        invitation_b = send_invitation(initiator=self.tutor, tutor_id=self.tutor.pk, tutee_id=tutee_b.pk)
        respond_to_invitation(invitation_id=invitation_a.pk, responder=self.tutee, accept=True)
        respond_to_invitation(invitation_id=invitation_b.pk, responder=tutee_b, accept=True)
        with self.assertRaises(ValidationError):
            create_admin_pairing(
                admin=admin, tutor_id=self.tutor.pk, tutee_id=tutee_c.pk, semester_id=self.semester.pk,
            )
        self.assertEqual(Pairing.objects.filter(tutor=self.tutor, status=PairingStatus.ACTIVE).count(), 2)

    def test_admin_pairing_still_enforces_program_roster_and_existing_tutor_checks(self):
        admin = User.objects.create_superuser(username="PAIR-ADMIN4", password="Admin-password-2026")
        with self.assertRaises(ValidationError):
            create_admin_pairing(
                admin=admin, tutor_id=self.tutor.pk, tutee_id=self.maryland.pk, semester_id=self.semester.pk,
            )
        create_admin_pairing(
            admin=admin, tutor_id=self.tutor.pk, tutee_id=self.tutee.pk, semester_id=self.semester.pk,
        )
        other_tutor = self.make_tutor("TUTOR320", "另一位老師", "Other Tutor")
        with self.assertRaises(ValidationError):
            create_admin_pairing(
                admin=admin, tutor_id=other_tutor.pk, tutee_id=self.tutee.pk, semester_id=self.semester.pk,
            )

    def test_non_admin_cannot_reach_create_pairing_view(self):
        self.client.force_login(self.tutor)
        response = self.client.post(reverse("tutoring:create_pairing"), {
            "tutor": self.tutor.pk, "tutee": self.tutee.pk, "semester": self.semester.pk,
        })
        self.assertEqual(response.status_code, 404)
        self.assertFalse(Pairing.objects.filter(tutor=self.tutor, tutee=self.tutee).exists())


class ClassWorkflowTests(TestCase):
    def setUp(self):
        today = timezone.localdate()
        self.semester = Semester.objects.create(
            name_zh="V2 測試學期", name_en="V2 test semester",
            starts_on=today - timedelta(days=7), ends_on=today + timedelta(days=90),
            is_active=True,
        )
        self.ntnu_program = PartnerProgram.objects.get(code="NTNU")
        tutor_roster = RosterEntry.objects.create(
            student_id="CLASS-TUTOR", name_zh="課程老師", name_en="Class Tutor", role=Role.TUTOR,
            education_level=EducationLevel.MASTER, identity_category=IdentityCategory.LOCAL,
        )
        tutee_roster = RosterEntry.objects.create(
            student_id="CLASS-TUTEE", name_zh="課程學生", name_en="Class Student", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.ntnu_program,
        )
        self.tutor = User.objects.create_user(
            username="CLASS-TUTOR", password="Test-password-2026", role=Role.TUTOR,
            roster_entry=tutor_roster, name_zh="課程老師", name_en="Class Tutor",
        )
        self.tutee = User.objects.create_user(
            username="CLASS-TUTEE", password="Test-password-2026", role=Role.TUTEE,
            roster_entry=tutee_roster, name_zh="課程學生", name_en="Class Student",
        )
        self.pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)

    def aware(self, day, clock):
        return timezone.make_aware(datetime.combine(day, clock), timezone.get_current_timezone())

    def record_data(self, topic):
        return {
            "location": "綜合大樓 / General Building",
            "topic": topic,
            "content": "會話與發音練習",
            "reflection": "完成本次練習並互相回饋",
            "materials_used": "教科書第三課、口說練習卡",
            "individual_progress": "發音有進步，句子長度可再增加",
            "remarks": "",
        }

    def test_schedule_class_view_returns_to_schedule_tab_on_success_and_error(self):
        """schedule_class's form lives on the dashboard's #schedule tab; both the success
        and validation-error redirects used to drop the tutor back on #overview instead."""
        self.client.force_login(self.tutor)
        class_date = timezone.localdate() + timedelta(days=1)
        response = self.client.post(
            reverse("tutoring:schedule_class"),
            {
                "pairing": self.pairing.pk,
                "class_date": class_date.isoformat(),
                "start_time": "10:00",
                "duration": "1.0",
            },
        )
        self.assertRedirects(response, reverse("accounts:dashboard") + "#schedule")

        invalid_response = self.client.post(reverse("tutoring:schedule_class"), {})
        self.assertRedirects(invalid_response, reverse("accounts:dashboard") + "#schedule")

    def test_class_cancel_returns_to_schedule_tab(self):
        class_date = timezone.localdate() + timedelta(days=1)
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date, start_time=time(10), duration="1.0"
        )[0]
        self.client.force_login(self.tutor)
        response = self.client.post(reverse("tutoring:class_cancel", args=[session.pk]))
        self.assertRedirects(response, reverse("accounts:dashboard") + "#schedule")

    def test_makeup_review_without_next_param_returns_to_makeup_review_tab(self):
        class_date = timezone.localdate() + timedelta(days=1)
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date, start_time=time(10), duration="1.0"
        )[0]
        MakeupReview.objects.create(session=session, status=MakeupReviewStatus.PENDING)
        admin = User.objects.create_superuser(username="CLASS-MAKEUP-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(
            reverse("tutoring:makeup_review", args=[session.pk]), {"action": "approve"}
        )
        self.assertRedirects(response, reverse("accounts:dashboard") + "#makeup-review")

    def test_schedule_reserves_weekly_quota_and_dashboard_shows_class(self):
        # Anchor to the Tuesday/Wednesday of a future week instead of "today + 1/+2 days":
        # the old relative offsets silently crossed two different Mon-Sun weeks whenever
        # the test happened to run on a Saturday (today+1 = Sunday, today+2 = Monday), so
        # the weekly quota check below never actually triggered on that one day of the week.
        today = timezone.localdate()
        next_monday = today + timedelta(days=(7 - today.weekday()) % 7 or 7)
        class_date = next_monday + timedelta(days=1)
        sessions = schedule_classes(
            tutor=self.tutor,
            pairing=self.pairing,
            class_date=class_date,
            start_time=time(22, 30),
            duration="1.5",
        )
        self.assertEqual(len(sessions), 1)
        self.assertEqual(sessions[0].ends_at.date(), class_date + timedelta(days=1))
        with self.assertRaises(ValidationError):
            schedule_classes(
                tutor=self.tutor,
                pairing=self.pairing,
                class_date=class_date + timedelta(days=1),
                start_time=time(10),
                duration="1.0",
            )
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "我的課表")
        self.assertContains(response, reverse("tutoring:class_detail", args=[sessions[0].pk]))

    def test_start_time_accepts_five_minute_increments(self):
        class_date = timezone.localdate() + timedelta(days=1)
        session = schedule_classes(
            tutor=self.tutor,
            pairing=self.pairing,
            class_date=class_date,
            start_time=time(17, 20),
            duration="1.0",
        )[0]
        self.assertEqual(session.start_time, time(17, 20))
        with self.assertRaises(ValidationError):
            schedule_classes(
                tutor=self.tutor,
                pairing=self.pairing,
                class_date=class_date + timedelta(days=7),
                start_time=time(17, 22),
                duration="1.0",
            )

    def test_schedule_form_lists_only_five_minute_values(self):
        class_date = timezone.localdate() + timedelta(days=1)
        form = ScheduleClassForm(
            {
                "pairing": self.pairing.pk,
                "class_date": class_date.isoformat(),
                "start_time_0": "17",
                "start_time_1": "20",
                "duration": "1",
                "repeat_weekly": "",
                "repeat_until": "",
            },
            tutor=self.tutor,
        )
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["start_time"], time(17, 20))
        minute_choices = [value for value, _label in form.fields["start_time"].widget.widgets[1].choices]
        self.assertEqual(minute_choices, [f"{minute:02d}" for minute in range(0, 60, 5)])

    def test_both_records_and_mutual_confirmation_create_valid_hours(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        normal_now = self.aware(class_date, time(10, 30))
        check_in(session_id=session.pk, participant=self.tutor, now=normal_now)
        check_in(session_id=session.pk, participant=self.tutee, now=normal_now)
        submit_class_record(
            session_id=session.pk, author=self.tutor, data=self.record_data("老師紀錄"), now=normal_now
        )
        submit_class_record(
            session_id=session.pk, author=self.tutee, data=self.record_data("學生紀錄"), now=normal_now
        )
        confirm_counterpart(
            session_id=session.pk, reviewer=self.tutor, status=ConfirmationStatus.CONFIRMED
        )
        confirm_counterpart(
            session_id=session.pk, reviewer=self.tutee, status=ConfirmationStatus.CONFIRMED
        )
        session.refresh_from_db()
        self.assertTrue(class_is_valid(session))

    def test_class_record_materials_used_and_individual_progress_saved_and_shown_to_counterpart_and_admin(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        normal_now = self.aware(class_date, time(10, 30))
        check_in(session_id=session.pk, participant=self.tutor, now=normal_now)
        submit_class_record(
            session_id=session.pk,
            author=self.tutor,
            data={
                **self.record_data("聽說練習"),
                "materials_used": "課本第五課、圖卡教具",
                "individual_progress": "口說流暢度明顯提升，仍需加強聲調準確度",
            },
            now=normal_now,
        )
        record = ClassRecord.objects.get(session=session, author=self.tutor)
        self.assertEqual(record.materials_used, "課本第五課、圖卡教具")
        self.assertEqual(record.individual_progress, "口說流暢度明顯提升，仍需加強聲調準確度")

        self.client.force_login(self.tutee)
        response = self.client.get(reverse("tutoring:class_detail", args=[session.pk]))
        self.assertContains(response, "課本第五課、圖卡教具")
        self.assertContains(response, "口說流暢度明顯提升，仍需加強聲調準確度")

        admin = User.objects.create_superuser(username="RECORD-MATERIALS-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("tutoring:class_detail", args=[session.pk]))
        self.assertContains(response, "課本第五課、圖卡教具")
        self.assertContains(response, "口說流暢度明顯提升，仍需加強聲調準確度")

    def test_class_record_requires_materials_used_and_individual_progress(self):
        data = {**self.record_data("必填欄位測試")}
        del data["materials_used"]
        del data["individual_progress"]
        form = ClassRecordForm(data=data)
        self.assertFalse(form.is_valid())
        self.assertIn("materials_used", form.errors)
        self.assertIn("individual_progress", form.errors)

    def test_class_record_materials_used_enforces_200_char_limit(self):
        data = {
            **self.record_data("使用之教材上限測試"), "materials_used": "材" * 200,
            "evidence_links": ["https://example.com/evidence"],
        }
        form = ClassRecordForm(data=data)
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.fields["materials_used"].widget.attrs["maxlength"], "200")

        over_limit_data = {
            **self.record_data("使用之教材超過上限測試"), "materials_used": "材" * 201,
            "evidence_links": ["https://example.com/evidence"],
        }
        over_limit_form = ClassRecordForm(data=over_limit_data)
        self.assertFalse(over_limit_form.is_valid())
        self.assertIn("materials_used", over_limit_form.errors)

    def test_class_record_individual_progress_enforces_500_char_limit(self):
        data = {
            **self.record_data("個別學習情形上限測試"), "individual_progress": "況" * 500,
            "evidence_links": ["https://example.com/evidence"],
        }
        form = ClassRecordForm(data=data)
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.fields["individual_progress"].widget.attrs["maxlength"], "500")

        over_limit_data = {
            **self.record_data("個別學習情形超過上限測試"), "individual_progress": "況" * 501,
            "evidence_links": ["https://example.com/evidence"],
        }
        over_limit_form = ClassRecordForm(data=over_limit_data)
        self.assertFalse(over_limit_form.is_valid())
        self.assertIn("individual_progress", over_limit_form.errors)

    def test_class_record_content_and_remarks_enforce_500_char_limit(self):
        base = {
            "location": "綜合大樓 / General Building", "topic": "課堂主題",
            "materials_used": "課本、教具", "individual_progress": "進度正常",
            "evidence_links": ["https://example.com/evidence"],
        }
        at_limit_data = {**base, "content": "內" * 500, "remarks": "備" * 500}
        at_limit_form = ClassRecordForm(data=at_limit_data)
        self.assertTrue(at_limit_form.is_valid(), at_limit_form.errors)
        self.assertEqual(at_limit_form.fields["content"].widget.attrs["maxlength"], "500")
        self.assertEqual(at_limit_form.fields["content"].widget.attrs["data-character-count"], "500")

        over_limit_data = {**base, "content": "內" * 501, "remarks": ""}
        over_limit_form = ClassRecordForm(data=over_limit_data)
        self.assertFalse(over_limit_form.is_valid())
        self.assertIn("content", over_limit_form.errors)

        over_limit_remarks_data = {**base, "content": "課堂內容", "remarks": "備" * 501}
        over_limit_remarks_form = ClassRecordForm(data=over_limit_remarks_data)
        self.assertFalse(over_limit_remarks_form.is_valid())
        self.assertIn("remarks", over_limit_remarks_form.errors)

    def test_class_record_attachment_saved_and_downloadable_by_counterpart_and_admin(self):
        """This exercises a Tutor record with an attachment via the service layer directly
        (bypassing ClassRecordForm, which no longer offers Tutors an attachment field after
        item 14) — representing pre-item-14 legacy data. class_detail/admin_record_card
        fall back to showing the attachment whenever evidence_links is empty, so this old
        shape must stay visible."""
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        normal_now = self.aware(class_date, time(10, 30))
        check_in(session_id=session.pk, participant=self.tutor, now=normal_now)
        submit_class_record(
            session_id=session.pk,
            author=self.tutor,
            data={
                **self.record_data("附件測試"),
                "attachment": SimpleUploadedFile("proof.pdf", b"%PDF-1.4 test", content_type="application/pdf"),
            },
            now=normal_now,
        )
        record = ClassRecord.objects.get(session=session, author=self.tutor)
        self.assertTrue(record.attachment.name)
        # Item 3 (batch 3): the stored path is now a randomized UUID name, not the
        # original filename, so attachment_filename must come from the separately
        # tracked original_attachment_filename instead of the storage path.
        self.assertEqual(record.attachment_filename, "proof.pdf")
        self.assertNotEqual(Path(record.attachment.name).name, "proof.pdf")

        download_url = reverse("tutoring:download_class_record_attachment", args=[record.pk])

        self.client.force_login(self.tutee)
        response = self.client.get(reverse("tutoring:class_detail", args=[session.pk]))
        self.assertContains(response, download_url)
        self.assertNotContains(response, record.attachment.url)

        admin = User.objects.create_superuser(username="RECORD-ATTACHMENT-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("tutoring:class_detail", args=[session.pk]))
        self.assertContains(response, download_url)
        self.assertNotContains(response, record.attachment.url)

    def make_record_with_attachment(self):
        """Shared setup for download_class_record_attachment authorization tests below."""
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        now = self.aware(class_date, time(10, 30))
        check_in(session_id=session.pk, participant=self.tutee, now=now)
        submit_class_record(
            session_id=session.pk,
            author=self.tutee,
            data={
                **self.record_data("附件下載測試"),
                "attachment": SimpleUploadedFile("evidence.pdf", b"%PDF-1.4 test", content_type="application/pdf"),
            },
            now=now,
        )
        record = ClassRecord.objects.get(session=session, author=self.tutee)
        return session, record

    def test_participant_can_download_attachment_with_private_headers(self):
        _, record = self.make_record_with_attachment()
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("tutoring:download_class_record_attachment", args=[record.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("evidence.pdf", response["Content-Disposition"])
        self.assertEqual(response["Cache-Control"], "private, no-store")
        self.assertEqual(response["X-Content-Type-Options"], "nosniff")

    def test_admin_can_download_attachment(self):
        _, record = self.make_record_with_attachment()
        admin = User.objects.create_superuser(username="DOWNLOAD-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("tutoring:download_class_record_attachment", args=[record.pk]))
        self.assertEqual(response.status_code, 200)

    def test_unrelated_user_cannot_download_attachment(self):
        _, record = self.make_record_with_attachment()
        outsider = User.objects.create_user(
            username="OUTSIDER-TUTOR", password="Test-password-2026", role=Role.TUTOR,
        )
        self.client.force_login(outsider)
        response = self.client.get(reverse("tutoring:download_class_record_attachment", args=[record.pk]))
        self.assertEqual(response.status_code, 404)

    def test_unauthenticated_user_cannot_download_attachment(self):
        _, record = self.make_record_with_attachment()
        response = self.client.get(reverse("tutoring:download_class_record_attachment", args=[record.pk]))
        self.assertNotEqual(response.status_code, 200)

    def test_record_without_attachment_is_not_downloadable(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        now = self.aware(class_date, time(10, 30))
        check_in(session_id=session.pk, participant=self.tutee, now=now)
        submit_class_record(session_id=session.pk, author=self.tutee, data=self.record_data("無附件"), now=now)
        record = ClassRecord.objects.get(session=session, author=self.tutee)
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("tutoring:download_class_record_attachment", args=[record.pk]))
        self.assertEqual(response.status_code, 404)

    def test_participant_can_still_download_attachment_after_pairing_ends(self):
        """Item 9 (batch 3): download access mirrors class_detail's own access rule
        (_session_for_user), which doesn't gate on Pairing status — so this is already
        consistent with the existing "read-only history after pairing ends" behavior
        documented for messaging (CLAUDE.md 4.8), not a new policy decision."""
        _, record = self.make_record_with_attachment()
        self.pairing.status = PairingStatus.ENDED
        self.pairing.save()
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("tutoring:download_class_record_attachment", args=[record.pk]))
        self.assertEqual(response.status_code, 200)

    def test_both_participants_use_evidence_links_instead_of_attachments(self):
        tutor_form = ClassRecordForm(author=self.tutor)
        self.assertNotIn("attachment", tutor_form.fields)
        self.assertIn("evidence_links", tutor_form.fields)
        self.assertTrue(tutor_form.fields["evidence_links"].required)
        self.assertIn("當次上課佐證連結", tutor_form.fields["evidence_links"].help_text)
        self.assertIn("共用檢視權限", tutor_form.fields["evidence_links"].help_text)
        self.assertIn("administrators", tutor_form.fields["evidence_links"].help_text)
        self.assertIn('class="evidence-help-points"', tutor_form.fields["evidence_links"].help_text)

        tutee_form = ClassRecordForm(author=self.tutee)
        self.assertNotIn("attachment", tutee_form.fields)
        self.assertIn("evidence_links", tutee_form.fields)
        # Tutee's evidence links are optional; Tutor's remain required (see below).
        self.assertFalse(tutee_form.fields["evidence_links"].required)
        self.assertIn("選填", tutee_form.fields["evidence_links"].label)
        self.assertIn("共用檢視權限", tutee_form.fields["evidence_links"].help_text)

    def test_tutor_record_requires_at_least_one_evidence_link(self):
        data = self.record_data("佐證連結測試")
        form = ClassRecordForm(data=data, author=self.tutor)
        self.assertFalse(form.is_valid())
        self.assertIn("evidence_links", form.errors)

    def test_tutee_record_evidence_link_is_optional(self):
        data = self.record_data("學生佐證連結測試")
        form = ClassRecordForm(data=data, author=self.tutee)
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["evidence_links"], [])

    def test_tutee_record_still_capped_at_five_evidence_links(self):
        data = {
            **self.record_data("學生佐證連結上限測試"),
            "evidence_links": [f"https://drive.example.com/file{i}" for i in range(6)],
        }
        form = ClassRecordForm(data=data, author=self.tutee)
        self.assertFalse(form.is_valid())
        self.assertIn("最多只能提供 5 個佐證連結", str(form.errors["evidence_links"]))

    def test_class_detail_shows_optional_placeholder_for_tutee_skipped_evidence_links(self):
        """A tutee who legitimately skips the now-optional evidence links must not see the
        legacy "未上傳 / Not uploaded" attachment wording, which would wrongly imply they
        forgot a required upload."""
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        now = self.aware(class_date, time(10, 30))
        check_in(session_id=session.pk, participant=self.tutee, now=now)
        submit_class_record(session_id=session.pk, author=self.tutee, data=self.record_data("無佐證連結"), now=now)

        self.client.force_login(self.tutor)
        response = self.client.get(reverse("tutoring:class_detail", args=[session.pk]))
        self.assertContains(response, "未提供（選填）/ Not provided (optional)")
        self.assertNotContains(response, "未上傳 / Not uploaded")

        admin = User.objects.create_superuser(username="RECORD-EVIDENCE-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("tutoring:class_detail", args=[session.pk]))
        self.assertContains(response, "未提供（選填）/ Not provided (optional)")
        self.assertNotContains(response, "未上傳 / Not uploaded")

    def test_tutor_record_rejects_more_than_five_links(self):
        data = {
            **self.record_data("佐證連結測試"),
            "evidence_links": [f"https://drive.example.com/file{i}" for i in range(6)],
        }
        form = ClassRecordForm(data=data, author=self.tutor)
        self.assertFalse(form.is_valid())
        self.assertIn("最多只能提供 5 個佐證連結", str(form.errors["evidence_links"]))

    def test_tutor_record_rejects_non_https_links(self):
        data = {**self.record_data("佐證連結測試"), "evidence_links": ["http://drive.example.com/file"]}
        form = ClassRecordForm(data=data, author=self.tutor)
        self.assertFalse(form.is_valid())
        self.assertIn("不是合法的 https 網址", str(form.errors["evidence_links"]))

    def test_tutor_record_accepts_one_to_five_https_links_from_any_domain(self):
        """Item 14: accepts any https:// URL, not just Google Drive/YouTube."""
        data = {
            **self.record_data("佐證連結測試"),
            "evidence_links": ["https://drive.example.com/a", "https://not-drive-or-youtube.example.org/b"],
        }
        form = ClassRecordForm(data=data, author=self.tutor)
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(
            form.cleaned_data["evidence_links"],
            ["https://drive.example.com/a", "https://not-drive-or-youtube.example.org/b"],
        )

    def test_tutor_submits_evidence_links_in_entered_order_via_class_detail_view(self):
        # This posts through the real view (not the service layer directly, like the
        # other tests in this class), so submit_class_record() below uses real
        # timezone.now() with no override. Floor the start time to the current 5-minute
        # mark so "now >= starts_at" holds no matter when the test actually runs, while
        # scheduling it relative to an hour-earlier creation time keeps it "in the future"
        # at creation, same as every other test's fixed 09:00→10:00 pattern.
        real_now = timezone.now()
        local_now = timezone.localtime(real_now)
        class_date = local_now.date()
        start_time = time(local_now.hour, (local_now.minute // 5) * 5)
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=start_time, duration="1.0", now=real_now - timedelta(hours=1),
        )[0]
        check_in(session_id=session.pk, participant=self.tutor)
        self.client.force_login(self.tutor)
        links = ["https://drive.example.com/c", "https://youtube.example.com/watch", "https://third.example.net/x"]
        response = self.client.post(
            reverse("tutoring:class_detail", args=[session.pk]),
            {**self.record_data("課堂佐證"), "evidence_links": links},
        )
        self.assertRedirects(response, reverse("tutoring:class_detail", args=[session.pk]))
        record = ClassRecord.objects.get(session=session, author=self.tutor)
        self.assertEqual(record.evidence_links, links)

        self.client.force_login(self.tutee)
        response = self.client.get(reverse("tutoring:class_detail", args=[session.pk]))
        for link in links:
            self.assertContains(response, f'href="{link}" target="_blank" rel="noopener noreferrer"')

    def test_makeup_reason_fields_only_shown_when_overdue(self):
        today = timezone.localdate()
        future_session = ClassSession.objects.create(
            pairing=self.pairing, class_date=today + timedelta(days=1), start_time=time(10, 0),
            duration=1, created_by=self.tutor,
        )
        overdue_session = ClassSession.objects.create(
            pairing=self.pairing, class_date=today - timedelta(days=3), start_time=time(10, 0),
            duration=1, created_by=self.tutor,
        )
        self.client.force_login(self.tutor)

        response = self.client.get(reverse("tutoring:class_detail", args=[future_session.pk]))
        self.assertNotContains(response, "逾時補簽原因")
        self.assertNotContains(response, "逾時補登原因")
        self.assertContains(response, "確認簽到 / Check in")
        self.assertContains(response, "送出紀錄 / Submit record")
        self.assertNotContains(response, "補簽到 / Makeup check-in")
        self.assertNotContains(response, "補填課堂紀錄 / Makeup record")

        response = self.client.get(reverse("tutoring:class_detail", args=[overdue_session.pk]))
        self.assertContains(response, "逾時補簽原因")
        self.assertContains(response, "逾時補登原因")
        self.assertContains(response, "補簽到 / Makeup check-in")
        self.assertContains(response, "補填課堂紀錄 / Makeup record")
        self.assertNotContains(response, "確認簽到 / Check in")
        self.assertNotContains(response, "送出紀錄 / Submit record")

    def test_makeup_record_requires_mutual_confirmation_and_admin_approval(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(8), duration="1.0", now=self.aware(class_date, time(7)),
        )[0]
        check_time = self.aware(class_date, time(8, 30))
        check_in(session_id=session.pk, participant=self.tutor, now=check_time)
        check_in(session_id=session.pk, participant=self.tutee, now=check_time)
        late = self.aware(class_date + timedelta(days=2), time(9))
        submit_class_record(
            session_id=session.pk, author=self.tutor, data=self.record_data("老師補登"),
            reason="忘記在期限內填寫", now=late,
        )
        submit_class_record(
            session_id=session.pk, author=self.tutee, data=self.record_data("學生補登"),
            reason="忘記在期限內填寫", now=late,
        )
        confirm_counterpart(session_id=session.pk, reviewer=self.tutor, status=ConfirmationStatus.CONFIRMED)
        confirm_counterpart(session_id=session.pk, reviewer=self.tutee, status=ConfirmationStatus.CONFIRMED)
        session.makeup_review.refresh_from_db()
        self.assertEqual(session.makeup_review.status, MakeupReviewStatus.PENDING)
        self.assertFalse(class_is_valid(session))
        admin = User.objects.create_superuser(username="CLASS-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        detail = self.client.get(reverse("tutoring:class_detail", args=[session.pk]))
        self.assertEqual(detail.status_code, 200)
        self.assertContains(detail, "補登詳情")
        self.assertContains(detail, "老師補登")
        self.assertContains(detail, "忘記在期限內填寫")
        # Once both parties confirm, the review moves to PENDING and admins should see
        # "waiting for admin approval" wording, not the earlier "waiting for mutual
        # confirmation" state or a generic "pending review" label.
        self.assertContains(detail, "等待管理員核准")
        dashboard = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(dashboard, "等待管理員核准")
        review_makeup(session_id=session.pk, admin=admin, approve=True)
        session.makeup_review.refresh_from_db()
        self.assertTrue(class_is_valid(session))
        history = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(history, "已核准")
        self.assertContains(history, "補課堂紀錄")

    def test_tutor_and_tutee_schedule_badge_reflects_makeup_review_status_not_generic_waiting(self):
        """Tutor/Tutee's own class list (class_schedule_group.html / class_history_list.html)
        computes its status badge from is_official/my_record/my_attendance alone, without
        looking at the actual MakeupReview status. Once both parties confirm a makeup class,
        it should show the review's real "等待管理員核准" state, not the generic "等待雙方完成
        / Waiting" text that never changes even after admin approval is the only thing left."""
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(8), duration="1.0", now=self.aware(class_date, time(7)),
        )[0]
        check_time = self.aware(class_date, time(8, 30))
        check_in(session_id=session.pk, participant=self.tutor, now=check_time)
        check_in(session_id=session.pk, participant=self.tutee, now=check_time)
        late = self.aware(class_date + timedelta(days=2), time(9))
        submit_class_record(
            session_id=session.pk, author=self.tutor, data=self.record_data("老師補登"),
            reason="忘記在期限內填寫", now=late,
        )
        submit_class_record(
            session_id=session.pk, author=self.tutee, data=self.record_data("學生補登"),
            reason="忘記在期限內填寫", now=late,
        )

        self.client.force_login(self.tutor)
        before = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(before, "等待雙方完成 / Waiting")

        confirm_counterpart(session_id=session.pk, reviewer=self.tutor, status=ConfirmationStatus.CONFIRMED)
        confirm_counterpart(session_id=session.pk, reviewer=self.tutee, status=ConfirmationStatus.CONFIRMED)
        session.makeup_review.refresh_from_db()
        self.assertEqual(session.makeup_review.status, MakeupReviewStatus.PENDING)

        after_tutor = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(after_tutor, "等待管理員核准")
        self.assertNotContains(after_tutor, "等待雙方完成 / Waiting")

        self.client.force_login(self.tutee)
        after_tutee = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(after_tutee, "等待管理員核准")
        self.assertNotContains(after_tutee, "等待雙方完成 / Waiting")

        admin = User.objects.create_superuser(username="SCHEDULE-BADGE-ADMIN", password="Admin-password-2026")
        review_makeup(session_id=session.pk, admin=admin, approve=False, note="資料不完整")
        session.makeup_review.refresh_from_db()
        self.assertEqual(session.makeup_review.status, MakeupReviewStatus.REJECTED)
        self.client.force_login(self.tutor)
        after_reject = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(after_reject, "未核准 / Rejected")
        self.assertContains(after_reject, 'class-status rejected')
        self.assertNotContains(after_reject, "等待雙方完成 / Waiting")

    def test_schedule_keeps_future_classes_and_moves_past_classes_to_hours_history(self):
        today = timezone.localdate()
        past = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=today - timedelta(days=1),
            start_time=time(10), duration="0.5", now=self.aware(today - timedelta(days=1), time(9)),
        )[0]
        upcoming = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=today + timedelta(days=2),
            start_time=time(10), duration="0.5",
        )[0]
        future = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=today + timedelta(days=10),
            start_time=time(10), duration="0.5",
        )[0]
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "即將課程")
        self.assertContains(response, "未來課程")
        self.assertContains(response, "學期課程")
        self.assertNotContains(response, "schedule-group-past")
        self.assertContains(response, "semester-hours-card")
        self.assertContains(response, "累積時數總覽")
        self.assertContains(response, "data-info-toggle")
        self.assertEqual([row.pk for row in response.context["upcoming_sessions"]], [upcoming.pk])
        self.assertEqual([row.pk for row in response.context["future_sessions"]], [future.pk])
        self.assertEqual([row.pk for row in response.context["past_sessions"]], [past.pk])

    def test_recent_missed_class_can_move_to_future_and_releases_old_week(self):
        today = timezone.localdate()
        old_date = today - timedelta(days=1)
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=old_date,
            start_time=time(10), duration="2.0", now=self.aware(old_date, time(9)),
        )[0]
        new_date = today + timedelta(days=8)
        changed = reschedule_class(
            session_id=session.pk,
            tutor=self.tutor,
            class_date=new_date,
            start_time=time(23),
            duration="1.5",
            now=self.aware(today, time(12)),
        )
        self.assertEqual(changed[0].class_date, new_date)
        self.assertEqual(changed[0].duration, 1.5)

    def test_class_alert_appears_for_admin_and_disappears_after_reporter_cancels(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        with self.assertRaises(ValidationError):
            report_class_alert(
                session_id=session.pk,
                reporter=self.tutor,
                reason=ClassAlertReason.CANNOT_REACH,
                now=self.aware(class_date, time(9, 55)),
            )
        alert = report_class_alert(
            session_id=session.pk,
            reporter=self.tutor,
            reason=ClassAlertReason.CANNOT_REACH,
            now=self.aware(class_date, time(10, 15)),
        )
        self.assertEqual(alert.subject, self.tutee)
        admin = User.objects.create_superuser(username="ALERT-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "課堂通報")
        self.assertContains(response, "聯絡不到對方")
        cancel_class_alert(alert_id=alert.pk, reporter=self.tutor)
        alert.refresh_from_db()
        self.assertEqual(alert.status, ClassAlertStatus.CANCELLED)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertNotContains(response, "聯絡不到對方")

    def test_admin_can_resolve_class_alert_and_history_shows_note(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        alert = report_class_alert(
            session_id=session.pk,
            reporter=self.tutor,
            reason=ClassAlertReason.CANNOT_REACH,
            now=self.aware(class_date, time(10, 15)),
        )
        admin = User.objects.create_superuser(username="ALERT-RESOLVE-ADMIN", password="Admin-password-2026")
        resolved = resolve_class_alert(alert_id=alert.pk, admin=admin, note="已與雙方確認過情況")
        self.assertEqual(resolved.status, ClassAlertStatus.RESOLVED)
        self.assertEqual(resolved.resolved_by, admin)

        self.client.force_login(admin)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "已與雙方確認過情況")
        self.assertContains(response, "目前沒有待處理的課堂通報")

    def test_resolved_class_alert_cannot_be_cancelled_and_active_cannot_be_resolved_twice(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        alert = report_class_alert(
            session_id=session.pk,
            reporter=self.tutor,
            reason=ClassAlertReason.OTHER,
            note="其他狀況",
            now=self.aware(class_date, time(10, 15)),
        )
        admin = User.objects.create_superuser(username="ALERT-RESOLVE-ADMIN2", password="Admin-password-2026")
        resolve_class_alert(alert_id=alert.pk, admin=admin, note="")
        with self.assertRaises(ValidationError):
            cancel_class_alert(alert_id=alert.pk, reporter=self.tutor)
        with self.assertRaises(ValidationError):
            resolve_class_alert(alert_id=alert.pk, admin=admin, note="")

    def test_non_admin_cannot_resolve_class_alert(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        alert = report_class_alert(
            session_id=session.pk,
            reporter=self.tutor,
            reason=ClassAlertReason.OTHER,
            note="其他狀況",
            now=self.aware(class_date, time(10, 15)),
        )
        with self.assertRaises(ValidationError):
            resolve_class_alert(alert_id=alert.pk, admin=self.tutee, note="")
        self.client.force_login(self.tutor)
        response = self.client.post(reverse("tutoring:resolve_alert", args=[alert.pk]), {"note": ""})
        self.assertEqual(response.status_code, 404)

    def test_incident_report_is_not_restricted_to_class_time_window(self):
        class_date = timezone.localdate() - timedelta(days=3)
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        report = submit_incident_report(
            session_id=session.pk,
            reporter=self.tutor,
            category=IncidentReportCategory.STUDENT_ABSENT,
            content="學生當天未出席，也聯絡不上。",
        )
        self.assertEqual(report.status, IncidentReportStatus.PENDING)
        self.assertEqual(report.reporter, self.tutor)

    def test_incident_report_rejects_non_participant(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        bystander = User.objects.create_user(
            username="BYSTANDER", password="Test-password-2026", role=Role.TUTOR
        )
        with self.assertRaises(ValidationError):
            submit_incident_report(
                session_id=session.pk,
                reporter=bystander,
                category=IncidentReportCategory.OTHER,
                content="不是這堂課的參與者。",
            )

    def test_incident_report_requires_valid_category_and_content(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        with self.assertRaises(ValidationError):
            submit_incident_report(
                session_id=session.pk, reporter=self.tutor, category="NOT_A_CATEGORY", content="測試"
            )
        with self.assertRaises(ValidationError):
            submit_incident_report(
                session_id=session.pk,
                reporter=self.tutor,
                category=IncidentReportCategory.OTHER,
                content="   ",
            )

    def test_admin_can_resolve_incident_report_and_dashboard_history_updates(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        report = submit_incident_report(
            session_id=session.pk,
            reporter=self.tutee,
            category=IncidentReportCategory.VENUE_ISSUE,
            content="教室臨時被佔用。",
        )
        admin = User.objects.create_superuser(username="INCIDENT-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "異常回報")
        self.assertContains(response, "教室臨時被佔用")

        response = self.client.post(
            reverse("tutoring:resolve_incident_report", args=[report.pk]),
            {"note": "已協調改到 202 教室"},
        )
        self.assertRedirects(response, reverse("accounts:dashboard") + "#incident-reports")
        report.refresh_from_db()
        self.assertEqual(report.status, IncidentReportStatus.RESOLVED)
        self.assertEqual(report.resolved_by, admin)
        self.assertEqual(report.resolution_note, "已協調改到 202 教室")

        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "已協調改到 202 教室")

    def test_non_admin_cannot_resolve_incident_report(self):
        class_date = timezone.localdate()
        session = schedule_classes(
            tutor=self.tutor, pairing=self.pairing, class_date=class_date,
            start_time=time(10), duration="1.0", now=self.aware(class_date, time(9)),
        )[0]
        report = submit_incident_report(
            session_id=session.pk,
            reporter=self.tutor,
            category=IncidentReportCategory.OTHER,
            content="測試內容",
        )
        with self.assertRaises(ValidationError):
            resolve_incident_report(report_id=report.pk, admin=self.tutee, note="")
        self.client.force_login(self.tutor)
        response = self.client.post(
            reverse("tutoring:resolve_incident_report", args=[report.pk]), {"note": ""}
        )
        self.assertEqual(response.status_code, 404)


class V2FeatureTests(TestCase):
    def setUp(self):
        today = timezone.localdate()
        self.semester = Semester.objects.create(
            name_zh="目前學期", name_en="Current semester",
            starts_on=today - timedelta(days=20), ends_on=today + timedelta(days=20),
            is_active=True,
        )
        self.ntnu_program = PartnerProgram.objects.get(code="NTNU")
        tutor_roster = RosterEntry.objects.create(
            student_id="V2-TUTOR", name_zh="V2老師", role=Role.TUTOR,
            education_level=EducationLevel.MASTER, identity_category=IdentityCategory.LOCAL,
        )
        tutee_roster = RosterEntry.objects.create(
            student_id="V2-TUTEE", name_zh="V2學生", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.ntnu_program,
        )
        self.tutor = User.objects.create_user(username="V2-TUTOR", password="Password-2026", role=Role.TUTOR, roster_entry=tutor_roster)
        self.tutee = User.objects.create_user(username="V2-TUTEE", password="Password-2026", role=Role.TUTEE, roster_entry=tutee_roster)
        self.pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)

    def test_active_pairing_can_message_and_ended_pairing_is_read_only(self):
        self.client.force_login(self.tutor)
        url = reverse("tutoring:pairing_messages", args=[self.pairing.pk])
        response = self.client.post(url, {"body": "明天見 / See you tomorrow"})
        self.assertRedirects(response, url)
        self.assertTrue(PairingMessage.objects.filter(pairing=self.pairing, sender=self.tutor).exists())
        self.pairing.status = PairingStatus.ENDED
        self.pairing.save(update_fields=["status"])
        response = self.client.post(url, {"body": "Should not send"}, follow=True)
        self.assertContains(response, "只能查看歷史訊息")
        self.assertEqual(PairingMessage.objects.filter(pairing=self.pairing).count(), 1)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "過往對話紀錄")
        self.assertContains(response, "查看紀錄 / View history")
        self.assertContains(response, url)
        response = self.client.get(url)
        self.assertContains(response, "明天見 / See you tomorrow")
        self.assertContains(response, "已結束 · 僅供查看")
        self.assertNotContains(response, reverse("accounts:matched_profile", args=[self.tutee.pk]))

    def test_messages_page_shows_counterparts_email_next_to_their_name(self):
        """Item 11: the messages page's counterpart name area shows their email."""
        self.tutee.email = "v2-tutee@example.com"
        self.tutee.save(update_fields=["email"])
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("tutoring:pairing_messages", args=[self.pairing.pk]))
        self.assertContains(response, "（v2-tutee@example.com）")
        self.assertContains(response, '<small class="conversation-person-email">（v2-tutee@example.com）</small>', html=False)

        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, '<small class="conversation-person-email">（v2-tutee@example.com）</small>', html=False)

    def test_dashboard_shows_unread_badge_and_preview_then_clears_on_open(self):
        self.client.force_login(self.tutee)
        self.client.post(
            reverse("tutoring:pairing_messages", args=[self.pairing.pk]),
            {"body": "老師好，請問明天上課地點在哪裡？"},
        )
        message_url = reverse("tutoring:pairing_messages", args=[self.pairing.pk])

        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(response, "老師好，請問明天上課地點在哪裡？")
        self.assertContains(response, "私訊<small>Messages</small></b><em>1</em>", html=False)
        self.assertContains(response, '<b class="unread-badge">1</b>', html=False)

        self.client.get(message_url)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertNotContains(response, 'unread-badge')

    def test_conversation_list_sorts_by_most_recent_message(self):
        other_tutee_roster = RosterEntry.objects.create(
            student_id="V2-TUTEE2", name_zh="V2學生二", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.ntnu_program,
        )
        other_tutee = User.objects.create_user(
            username="V2-TUTEE2", password="Password-2026", role=Role.TUTEE, roster_entry=other_tutee_roster,
        )
        other_pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=other_tutee)

        self.client.force_login(self.tutee)
        self.client.post(reverse("tutoring:pairing_messages", args=[self.pairing.pk]), {"body": "第一則訊息"})
        self.client.force_login(other_tutee)
        self.client.post(reverse("tutoring:pairing_messages", args=[other_pairing.pk]), {"body": "比較新的訊息"})

        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:dashboard"))
        content = response.content.decode()
        self.assertLess(content.index("比較新的訊息"), content.index("第一則訊息"))

    def test_unrelated_user_cannot_open_messages(self):
        outsider = User.objects.create_user(username="V2-OUT", password="Password-2026", role=Role.TUTEE)
        self.client.force_login(outsider)
        self.assertEqual(self.client.get(reverse("tutoring:pairing_messages", args=[self.pairing.pk])).status_code, 404)

    def test_admin_class_overview_links_to_tutor_schedule(self):
        ClassSession.objects.create(
            pairing=self.pairing,
            class_date=timezone.localdate(),
            start_time=time(10, 0),
            duration=1,
            created_by=self.tutor,
        )
        admin = User.objects.create_superuser(username="CLASS-OVERVIEW-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.get(reverse("accounts:dashboard"))
        schedule_url = reverse("accounts:admin_tutor_schedule", args=[self.tutor.pk])
        self.assertContains(response, "老師名單")
        self.assertContains(response, schedule_url)
        self.assertNotContains(response, "Latest 100 classes")
        response = self.client.get(f"{schedule_url}?semester={self.semester.pk}")
        self.assertContains(response, "TUTOR SCHEDULE")
        self.assertContains(response, self.tutee.username)

    def test_non_admin_cannot_open_admin_tutor_schedule(self):
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:admin_tutor_schedule", args=[self.tutor.pk]))
        self.assertEqual(response.status_code, 404)

    def test_admin_user_profile_aggregates_pairings_hours_and_reports(self):
        session = ClassSession.objects.create(
            pairing=self.pairing, class_date=timezone.localdate(), start_time=time(10, 0),
            duration=1, created_by=self.tutor,
        )
        ClassAlert.objects.create(
            session=session, reporter=self.tutor, subject=self.tutee, reason=ClassAlertReason.CANNOT_REACH,
        )
        IncidentReport.objects.create(
            session=session, reporter=self.tutee, category=IncidentReportCategory.VENUE_ISSUE, content="教室有問題",
        )
        admin = User.objects.create_superuser(username="PROFILE-ADMIN", password="Admin-password-2026")
        HourAdjustment.objects.create(
            user=self.tutor, semester=self.semester, program=self.ntnu_program,
            hours=Decimal("2.5"), reason="舊紙本資料補登測試", created_by=admin,
        )
        self.client.force_login(admin)

        response = self.client.get(reverse("accounts:admin_user_profile", args=[self.tutor.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.tutee.bilingual_name)
        self.assertContains(response, "場地問題 / Venue issue")
        self.assertContains(response, "舊紙本資料補登測試")
        self.assertContains(response, "+2.5 小時")

        response = self.client.get(reverse("accounts:admin_user_profile", args=[self.tutee.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.tutor.bilingual_name)
        self.assertContains(response, "場地問題 / Venue issue")

    def test_non_admin_cannot_open_admin_user_profile(self):
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("accounts:admin_user_profile", args=[self.tutee.pk]))
        self.assertEqual(response.status_code, 404)

    def test_admin_export_defaults_to_xlsx_and_can_filter_by_program_semester_and_specific_user(self):
        session = ClassSession.objects.create(
            pairing=self.pairing,
            class_date=timezone.localdate(),
            start_time=time(11, 0), duration=1, created_by=self.tutor,
        )
        admin = User.objects.create_superuser(username="EXPORT-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": self.ntnu_program.pk,
            "audience": "specific",
            "user_ids": [self.tutor.pk],
            "period_mode": "semester",
            "semester_id": self.semester.pk,
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertIn(str(session.class_date), rows[1])

    def test_admin_export_program_filter_excludes_same_tutors_other_program_classes(self):
        ntnu_session = ClassSession.objects.create(
            pairing=self.pairing,
            class_date=timezone.localdate(),
            start_time=time(11, 0), duration=1, created_by=self.tutor,
        )
        maryland = PartnerProgram.objects.get(code="MARYLAND")
        maryland_roster = RosterEntry.objects.create(
            student_id="EXPORT-MARYLAND", name_zh="馬里蘭學生", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE,
            identity_category=IdentityCategory.INTERNATIONAL, program=maryland,
        )
        maryland_tutee = User.objects.create_user(
            username="EXPORT-MARYLAND", password="Password-2026",
            role=Role.TUTEE, roster_entry=maryland_roster,
        )
        maryland_pairing = Pairing.objects.create(
            semester=self.semester, tutor=self.tutor, tutee=maryland_tutee,
        )
        ClassSession.objects.create(
            pairing=maryland_pairing, class_date=timezone.localdate(),
            start_time=time(12, 0), duration=1, created_by=self.tutor,
        )
        admin = User.objects.create_superuser(
            username="EXPORT-PROGRAM-ADMIN", password="Admin-password-2026"
        )
        self.client.force_login(admin)

        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": self.ntnu_program.pk,
            "audience": "specific",
            "user_ids": [self.tutor.pk],
            "period_mode": "semester",
            "semester_id": self.semester.pk,
            "file_format": "xlsx",
        })

        self.assertEqual(response.status_code, 200)
        rows = list(openpyxl.load_workbook(BytesIO(response.content)).active.iter_rows(values_only=True))
        exported = " ".join(str(value) for row in rows for value in row if value is not None)
        self.assertIn(str(ntnu_session.class_date), exported)
        self.assertIn(self.tutee.username, exported)
        self.assertNotIn(maryland_tutee.username, exported)

    def test_admin_export_can_select_all_students_in_program(self):
        ClassSession.objects.create(
            pairing=self.pairing, class_date=timezone.localdate(),
            start_time=time(11, 0), duration=1, created_by=self.tutor,
        )
        admin = User.objects.create_superuser(
            username="EXPORT-TUTEES-ADMIN", password="Admin-password-2026"
        )
        self.client.force_login(admin)

        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": self.ntnu_program.pk,
            "audience": "tutees",
            "period_mode": "semester",
            "semester_id": self.semester.pk,
            "file_format": "xlsx",
        })

        self.assertEqual(response.status_code, 200)
        rows = list(openpyxl.load_workbook(BytesIO(response.content)).active.iter_rows(values_only=True))
        self.assertEqual(rows[1][0], self.tutee.username)
        self.assertNotEqual(rows[1][0], self.tutor.username)

    def test_admin_export_rejects_specific_user_outside_selected_program(self):
        maryland = PartnerProgram.objects.get(code="MARYLAND")
        admin = User.objects.create_superuser(
            username="EXPORT-WRONG-PROGRAM-ADMIN", password="Admin-password-2026"
        )
        self.client.force_login(admin)

        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": maryland.pk,
            "audience": "specific",
            "user_ids": [self.tutee.pk],
            "period_mode": "semester",
            "semester_id": self.semester.pk,
        })

        self.assertRedirects(response, f"{reverse('accounts:dashboard')}#export")

    def test_admin_export_can_produce_real_xlsx(self):
        session = ClassSession.objects.create(
            pairing=self.pairing,
            class_date=timezone.localdate(),
            start_time=time(11, 0), duration=1, created_by=self.tutor,
        )
        admin = User.objects.create_superuser(username="EXPORT-XLSX-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": self.ntnu_program.pk,
            "audience": "specific",
            "user_ids": [self.tutor.pk],
            "period_mode": "semester",
            "semester_id": self.semester.pk,
            "file_format": "xlsx",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(response["Content-Disposition"].endswith('.xlsx"'))
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "學號 Student ID")
        self.assertIn(str(session.class_date), rows[1])

    def test_admin_export_can_select_output_fields(self):
        session = ClassSession.objects.create(
            pairing=self.pairing,
            class_date=timezone.localdate(),
            start_time=time(11, 0), duration=1, created_by=self.tutor,
        )
        admin = User.objects.create_superuser(username="EXPORT-FIELDS-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)

        dashboard = self.client.get(reverse("accounts:dashboard"))
        self.assertContains(dashboard, "選擇計畫")
        self.assertContains(dashboard, 'name="audience" value="tutors" checked', html=False)
        self.assertContains(dashboard, 'name="audience" value="tutees"', html=False)
        self.assertContains(dashboard, 'name="audience" value="specific"', html=False)
        self.assertContains(dashboard, "特定使用者")
        self.assertContains(dashboard, 'data-role="TUTOR"', html=False)
        self.assertContains(dashboard, 'data-role="TUTEE"', html=False)
        self.assertContains(dashboard, "選擇欄位")
        self.assertContains(dashboard, 'name="export_fields" value="date" checked', html=False)

        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": self.ntnu_program.pk,
            "audience": "specific",
            "user_ids": [self.tutor.pk],
            "period_mode": "semester",
            "semester_id": self.semester.pk,
            "field_selection_present": "1",
            "export_fields": ["date", "hours"],
            "file_format": "xlsx",
        })
        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0], ("日期 Date", "時數 Hours"))
        self.assertEqual(rows[1], (str(session.class_date), "1.0"))

    def test_admin_export_requires_at_least_one_selected_field(self):
        admin = User.objects.create_superuser(username="EXPORT-NO-FIELDS-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": self.ntnu_program.pk,
            "audience": "tutors",
            "period_mode": "semester",
            "semester_id": self.semester.pk,
            "field_selection_present": "1",
            "file_format": "xlsx",
        })
        self.assertRedirects(response, f"{reverse('accounts:dashboard')}#export")

    def test_admin_export_can_produce_csv(self):
        session = ClassSession.objects.create(
            pairing=self.pairing,
            class_date=timezone.localdate(),
            start_time=time(11, 0), duration=1, created_by=self.tutor,
        )
        admin = User.objects.create_superuser(username="EXPORT-CSV-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": self.ntnu_program.pk,
            "audience": "specific",
            "user_ids": [self.tutor.pk],
            "period_mode": "semester",
            "semester_id": self.semester.pk,
            "file_format": "csv",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertTrue(response["Content-Disposition"].endswith('.csv"'))
        decoded = response.content.decode("utf-8-sig")
        self.assertIn("學號 Student ID", decoded)
        self.assertIn(str(session.class_date), decoded)

    def test_admin_export_can_produce_pdf(self):
        ClassSession.objects.create(
            pairing=self.pairing,
            class_date=timezone.localdate(),
            start_time=time(11, 0), duration=1, created_by=self.tutor,
        )
        admin = User.objects.create_superuser(username="EXPORT-PDF-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": self.ntnu_program.pk,
            "audience": "specific",
            "user_ids": [self.tutor.pk],
            "period_mode": "semester",
            "semester_id": self.semester.pk,
            "file_format": "pdf",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response["Content-Disposition"].endswith('.pdf"'))
        self.assertTrue(response.content.startswith(b"%PDF"))
        from pypdf import PdfReader
        from pypdf.constants import UserAccessPermissions
        reader = PdfReader(BytesIO(response.content))
        self.assertGreaterEqual(len(reader.pages), 1)
        # Item 3: exported administrative reports get the same copy/extraction restriction
        # as certificates — see test_certificate_pdf_restricts_copy_but_allows_printing.
        self.assertTrue(reader.is_encrypted)
        permissions = reader.user_access_permissions
        self.assertTrue(permissions & UserAccessPermissions.PRINT)
        self.assertFalse(permissions & UserAccessPermissions.EXTRACT)

    def test_admin_export_rejects_reversed_date_range(self):
        admin = User.objects.create_superuser(username="EXPORT-RANGE-ADMIN", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(reverse("tutoring:export_excel"), {
            "program_id": self.ntnu_program.pk, "audience": "tutors", "period_mode": "range",
            "starts_on": "2026-08-01", "ends_on": "2026-07-01",
        })
        self.assertRedirects(response, f"{reverse('accounts:dashboard')}#export")

    def test_archived_past_semester_remains_downloadable(self):
        today = timezone.localdate()
        past = Semester.objects.create(
            name_zh="過去學期", name_en="Past semester",
            starts_on=today - timedelta(days=100), ends_on=today - timedelta(days=20),
            is_active=False,
        )
        form = HoursDownloadForm()
        self.assertIn(past, list(form.fields["semester"].queryset))

    def test_semester_makeup_and_download_deadlines(self):
        self.assertEqual(self.semester.makeup_deadline_at.date(), self.semester.ends_on + timedelta(days=1))
        self.assertEqual(self.semester.hours_download_at.date(), self.semester.ends_on + timedelta(days=3))

    def test_detailed_certificate_requires_at_least_one_field(self):
        today = timezone.localdate()
        form = HoursDownloadForm({
            "mode": "range", "starts_on": today - timedelta(days=200),
            "ends_on": today - timedelta(days=190), "version": "detailed",
        })
        self.assertFalse(form.is_valid())
        self.assertIn("detail_fields", form.errors)

    def test_summary_and_detailed_certificate_use_pdf_template(self):
        from pypdf import PdfReader

        today = timezone.localdate()
        data = {
            "user": self.tutor, "starts_on": today - timedelta(days=30), "ends_on": today,
            "sections": [], "total": 0, "generated_at": timezone.now(),
        }
        summary = build_hours_pdf(data, version="summary", program=self.ntnu_program)
        detailed = build_hours_pdf(data, version="detailed", detail_fields=["date", "hours"], program=self.ntnu_program)
        self.assertTrue(summary.startswith(b"%PDF"))
        self.assertTrue(detailed.startswith(b"%PDF"))
        self.assertEqual(len(PdfReader(BytesIO(summary)).pages), 1)
        self.assertIn("實習證明", PdfReader(BytesIO(summary)).pages[0].extract_text())


class PartnerProgramCertificateTests(TestCase):
    def setUp(self):
        today = timezone.localdate()
        self.semester = Semester.objects.create(
            name_zh="方案測試學期", name_en="Program test semester",
            starts_on=today - timedelta(days=20), ends_on=today - timedelta(days=10),
            is_active=False,
        )
        self.ntnu_program = PartnerProgram.objects.get(code="NTNU")
        self.maryland_program = PartnerProgram.objects.get(code="MARYLAND")
        tutor_roster = RosterEntry.objects.create(
            student_id="PP-TUTOR", name_zh="方案老師", role=Role.TUTOR,
            education_level=EducationLevel.MASTER, identity_category=IdentityCategory.LOCAL,
        )
        tutee_roster = RosterEntry.objects.create(
            student_id="PP-TUTEE", name_zh="方案學生", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.ntnu_program,
        )
        self.tutor = User.objects.create_user(
            username="PP-TUTOR", password="Password-2026", role=Role.TUTOR, roster_entry=tutor_roster
        )
        self.tutee = User.objects.create_user(
            username="PP-TUTEE", password="Password-2026", role=Role.TUTEE, roster_entry=tutee_roster
        )
        self.pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=self.tutee)
        ClassSession.objects.create(
            pairing=self.pairing, class_date=today - timedelta(days=15),
            start_time=time(10), duration=1, created_by=self.tutor,
        )

    def test_ntnu_tutee_can_now_download_hours(self):
        self.assertTrue(user_has_hour_records(self.tutee))

    def test_certificate_pdf_restricts_copy_but_allows_printing(self):
        """Item 3: best-effort "no copy/extraction" PDF permission. No open password is
        set (extract_text() below succeeds without ever calling reader.decrypt()), and
        printing stays allowed — only text/graphics extraction is denied."""
        from pypdf import PdfReader
        from pypdf.constants import UserAccessPermissions

        data = {
            "user": self.tutor, "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        content = build_hours_pdf(data, version="summary", program=self.ntnu_program)
        reader = PdfReader(BytesIO(content))
        self.assertTrue(reader.is_encrypted)
        self.assertTrue(reader.pages[0].extract_text().strip())
        permissions = reader.user_access_permissions
        self.assertTrue(permissions & UserAccessPermissions.PRINT)
        self.assertTrue(permissions & UserAccessPermissions.PRINT_TO_REPRESENTATION)
        self.assertFalse(permissions & UserAccessPermissions.EXTRACT)
        self.assertFalse(permissions & UserAccessPermissions.MODIFY)

    def test_tutor_available_programs_only_lists_programs_actually_tutored(self):
        codes = {program.code for program in tutor_available_programs(self.tutor)}
        self.assertEqual(codes, {"NTNU"})

        maryland_tutee_roster = RosterEntry.objects.create(
            student_id="PP-TUTEE-MD", name_zh="方案馬里蘭學生", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.maryland_program,
        )
        maryland_tutee = User.objects.create_user(
            username="PP-TUTEE-MD", password="Password-2026", role=Role.TUTEE, roster_entry=maryland_tutee_roster
        )
        maryland_pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=maryland_tutee)
        ClassSession.objects.create(
            pairing=maryland_pairing, class_date=timezone.localdate() - timedelta(days=15),
            start_time=time(14), duration=1, created_by=self.tutor,
        )
        codes = {program.code for program in tutor_available_programs(self.tutor)}
        self.assertEqual(codes, {"NTNU", "MARYLAND"})

    def test_program_selector_is_tutor_only_and_uses_practicum_label(self):
        tutor_form = HoursDownloadForm(user=self.tutor)
        self.assertEqual(tutor_form.fields["program"].label, "實習計劃 / Practicum program")
        self.assertNotIn("program", HoursDownloadForm(user=self.tutee).fields)

    def test_certificate_language_uses_select_menu(self):
        field = HoursDownloadForm(user=self.tutor).fields["language"]
        self.assertEqual(field.widget.__class__.__name__, "Select")
        self.assertEqual(list(field.choices), [("zh", "中文版 / Chinese"), ("en", "英文版 / English")])

    def test_download_hours_get_returns_to_hours_panel_after_login(self):
        self.client.force_login(self.tutor)
        response = self.client.get(reverse("tutoring:download_hours"))
        self.assertRedirects(response, f"{reverse('accounts:dashboard')}#hours")

    def test_download_hours_requires_program_for_tutor(self):
        self.client.force_login(self.tutor)
        today = timezone.localdate()
        response = self.client.post(
            reverse("tutoring:download_hours"),
            {
                "mode": "range", "starts_on": today - timedelta(days=10), "ends_on": today,
                "version": "summary", "language": "zh",
            },
        )
        self.assertRedirects(response, f"{reverse('accounts:dashboard')}#hours")
        messages = list(response.wsgi_request._messages)
        self.assertTrue(any("實習計劃" in str(message) for message in messages))

    def test_tutor_can_download_ntnu_certificate_via_dropdown(self):
        self.client.force_login(self.tutor)
        today = timezone.localdate()
        response = self.client.post(
            reverse("tutoring:download_hours"),
            {
                "mode": "range", "starts_on": today - timedelta(days=10), "ends_on": today,
                "version": "summary", "program": self.ntnu_program.pk, "language": "zh",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("attachment", response["Content-Disposition"])

    def test_tutor_can_preview_certificate_inline_without_forcing_download(self):
        self.client.force_login(self.tutor)
        today = timezone.localdate()
        response = self.client.post(
            reverse("tutoring:download_hours"),
            {
                "mode": "range", "starts_on": today - timedelta(days=10), "ends_on": today,
                "version": "summary", "program": self.ntnu_program.pk, "intent": "preview", "language": "zh",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("inline", response["Content-Disposition"])
        log = AuditLog.objects.get(event_type="HOURS_PDF_PREVIEWED")
        self.assertEqual(log.actor, self.tutor)

    def test_ntnu_tutee_downloads_tutee_specific_template(self):
        from pypdf import PdfReader

        self.client.force_login(self.tutee)
        today = timezone.localdate()
        response = self.client.post(
            reverse("tutoring:download_hours"),
            {
                "mode": "range", "starts_on": today - timedelta(days=10), "ends_on": today,
                "version": "summary", "language": "zh",
            },
        )
        self.assertEqual(response.status_code, 200)
        text = PdfReader(BytesIO(response.content)).pages[0].extract_text()
        self.assertIn("接受華語輔導", text)
        self.assertIn("受輔導證明", text)

    def test_ntnu_tutor_same_month_period_is_not_repeated(self):
        from pypdf import PdfReader

        data = {
            "user": self.tutor,
            "starts_on": date(2026, 7, 1),
            "ends_on": date(2026, 7, 31),
            "sections": [],
            "total": 0,
            "generated_at": timezone.now(),
        }
        content = build_hours_pdf(data, version="summary", program=self.ntnu_program)
        extracted = PdfReader(BytesIO(content)).pages[0].extract_text()
        compact_text = extracted.replace(" ", "")
        self.assertIn("民國115年7月", compact_text)
        self.assertNotIn("7月-7月", compact_text)
        self.assertIn("PP-TUTOR（學號：PP-TUTOR）\n", compact_text)
        self.assertNotIn("學號：PP-TUTOR\n）", compact_text)

    def test_tutor_can_download_maryland_certificate_now_that_template_is_shared(self):
        from pypdf import PdfReader

        maryland_tutee_roster = RosterEntry.objects.create(
            student_id="PP-TUTEE-MD2", name_zh="方案馬里蘭學生二", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.maryland_program,
        )
        maryland_tutee = User.objects.create_user(
            username="PP-TUTEE-MD2", password="Password-2026", role=Role.TUTEE, roster_entry=maryland_tutee_roster
        )
        maryland_pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=maryland_tutee)
        ClassSession.objects.create(
            pairing=maryland_pairing, class_date=timezone.localdate() - timedelta(days=15),
            start_time=time(14), duration=1, created_by=self.tutor,
        )
        self.client.force_login(self.tutor)
        today = timezone.localdate()
        response = self.client.post(
            reverse("tutoring:download_hours"),
            {
                "mode": "range", "starts_on": today - timedelta(days=10), "ends_on": today,
                "version": "summary", "program": self.maryland_program.pk, "language": "zh",
            },
        )
        self.assertEqual(response.status_code, 200)
        text = PdfReader(BytesIO(response.content)).pages[0].extract_text()
        self.assertIn("語言交換服務證明", text)

    def _make_verified_session(self, pairing, other_party, class_date):
        session = ClassSession.objects.create(
            pairing=pairing, class_date=class_date, start_time=time(10), duration=1, created_by=self.tutor,
        )
        signed_at = self.aware_datetime(class_date, time(10, 5))
        for participant in (self.tutor, other_party):
            Attendance.objects.create(session=session, participant=participant, signed_at=signed_at)
        for author in (self.tutor, other_party):
            ClassRecord.objects.create(
                session=session, author=author, location="測試教室", topic="會話練習",
                content="練習內容", reflection="回饋內容",
            )
        for reviewer, subject in ((self.tutor, other_party), (other_party, self.tutor)):
            ClassConfirmation.objects.create(
                session=session, reviewer=reviewer, subject=subject,
                attendance_confirmed=True, record_confirmed=True, status=ConfirmationStatus.CONFIRMED,
            )
        return session

    def aware_datetime(self, day, clock):
        return timezone.make_aware(datetime.combine(day, clock), timezone.get_current_timezone())

    def test_download_hours_only_counts_selected_programs_sessions(self):
        maryland_tutee_roster = RosterEntry.objects.create(
            student_id="PP-TUTEE-MD3", name_zh="方案馬里蘭學生三", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.maryland_program,
        )
        maryland_tutee = User.objects.create_user(
            username="PP-TUTEE-MD3", password="Password-2026", role=Role.TUTEE, roster_entry=maryland_tutee_roster
        )
        maryland_pairing = Pairing.objects.create(semester=self.semester, tutor=self.tutor, tutee=maryland_tutee)
        past_date = timezone.localdate() - timedelta(days=12)
        self._make_verified_session(self.pairing, self.tutee, past_date)
        self._make_verified_session(maryland_pairing, maryland_tutee, past_date)

        self.client.force_login(self.tutor)
        today = timezone.localdate()
        response = self.client.post(
            reverse("tutoring:download_hours"),
            {
                "mode": "range", "starts_on": today - timedelta(days=20), "ends_on": today,
                "version": "summary", "program": self.ntnu_program.pk, "language": "zh",
            },
        )
        self.assertEqual(response.status_code, 200)
        from pypdf import PdfReader
        text = PdfReader(BytesIO(response.content)).pages[0].extract_text()
        self.assertIn("總計授課 1 小時", text)

    def test_hour_adjustment_rejects_non_positive_hours_and_wrong_role(self):
        admin = User.objects.create_superuser(username="ADJ-ADMIN", password="Admin-password-2026")
        adjustment = HourAdjustment(
            user=self.tutor, semester=self.semester, program=self.ntnu_program,
            hours=Decimal("0"), reason="測試", created_by=admin,
        )
        with self.assertRaises(ValidationError):
            adjustment.full_clean()

        admin_as_subject = HourAdjustment(
            user=admin, semester=self.semester, program=self.ntnu_program,
            hours=Decimal("1"), reason="測試", created_by=admin,
        )
        with self.assertRaises(ValidationError):
            admin_as_subject.full_clean()

    def test_hour_adjustment_raises_certificate_total_only_for_matching_program_and_semester(self):
        admin = User.objects.create_superuser(username="ADJ-ADMIN2", password="Admin-password-2026")
        past_date = timezone.localdate() - timedelta(days=12)
        self._make_verified_session(self.pairing, self.tutee, past_date)
        HourAdjustment.objects.create(
            user=self.tutor, semester=self.semester, program=self.ntnu_program,
            hours=Decimal("2.5"), reason="舊紙本資料補登", created_by=admin,
        )
        HourAdjustment.objects.create(
            user=self.tutor, semester=self.semester, program=self.maryland_program,
            hours=Decimal("9"), reason="不同計畫，不應被算入", created_by=admin,
        )

        self.client.force_login(self.tutor)
        today = timezone.localdate()
        response = self.client.post(
            reverse("tutoring:download_hours"),
            {
                "mode": "range", "starts_on": today - timedelta(days=20), "ends_on": today,
                "version": "summary", "program": self.ntnu_program.pk, "language": "zh",
            },
        )
        self.assertEqual(response.status_code, 200)
        from pypdf import PdfReader
        text = PdfReader(BytesIO(response.content)).pages[0].extract_text()
        self.assertIn("總計授課 3.5 小時", text)

    def test_tutor_available_programs_includes_adjustment_only_program(self):
        admin = User.objects.create_superuser(username="ADJ-ADMIN3", password="Admin-password-2026")
        self.assertNotIn("MARYLAND", {program.code for program in tutor_available_programs(self.tutor)})
        HourAdjustment.objects.create(
            user=self.tutor, semester=self.semester, program=self.maryland_program,
            hours=Decimal("4"), reason="舊紙本資料，無資料庫課程紀錄", created_by=admin,
        )
        self.assertIn("MARYLAND", {program.code for program in tutor_available_programs(self.tutor)})

    def test_admin_creating_hour_adjustment_writes_audit_log(self):
        admin = User.objects.create_superuser(username="ADJ-ADMIN4", password="Admin-password-2026")
        self.client.force_login(admin)
        response = self.client.post(
            reverse("admin:tutoring_houradjustment_add"),
            {
                "user": self.tutor.pk, "semester": self.semester.pk, "program": self.ntnu_program.pk,
                "hours": "2.0", "reason": "舊紙本資料補登",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(HourAdjustment.objects.filter(user=self.tutor).count(), 1)
        adjustment = HourAdjustment.objects.get(user=self.tutor)
        self.assertEqual(adjustment.created_by, admin)
        self.assertTrue(
            AuditLog.objects.filter(event_type="HOUR_ADJUSTMENT_CREATED", target_user=self.tutor).exists()
        )

    def test_certificate_language_field_controls_filename_and_audit_log(self):
        self.client.force_login(self.tutor)
        today = timezone.localdate()
        response = self.client.post(
            reverse("tutoring:download_hours"),
            {
                "mode": "range", "starts_on": today - timedelta(days=10), "ends_on": today,
                "version": "summary", "program": self.ntnu_program.pk, "language": "en",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(f"-en-{today}.pdf", response["Content-Disposition"])
        log = AuditLog.objects.get(event_type="HOURS_PDF_DOWNLOADED", target_user=self.tutor)
        self.assertEqual(log.metadata["language"], "en")

    def test_english_certificate_uses_only_english_text_and_gregorian_date(self):
        from pypdf import PdfReader

        data = {
            "user": self.tutor, "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        content = build_hours_pdf(data, version="summary", program=self.ntnu_program, language="en")
        text = PdfReader(BytesIO(content)).pages[0].extract_text()
        self.assertIn("Certificate of Counseling Practicum", text)
        self.assertNotIn("實習證明", text)
        self.assertNotIn("民國", text)

    def test_certificate_issue_date_uses_taiwan_local_date(self):
        from pypdf import PdfReader

        data = {
            "user": self.tutor, "starts_on": date(2026, 7, 1), "ends_on": date(2026, 8, 31),
            "sections": [], "total": 2,
            # 16:30 UTC is already the following calendar day in Taiwan.
            "generated_at": datetime.fromisoformat("2026-08-10T16:30:00+00:00"),
        }
        content = build_hours_pdf(data, version="summary", program=self.ntnu_program, language="zh")
        compact_text = PdfReader(BytesIO(content)).pages[0].extract_text().replace(" ", "").replace("\n", "")
        self.assertIn("中華民國115年8月11日", compact_text)

    def test_detailed_certificate_reserves_footer_after_six_rows(self):
        from pypdf import PdfReader

        session = ClassSession.objects.filter(pairing=self.pairing).first()
        detail_row = {
            "session": session,
            "counterpart": self.tutee,
            "topic": "會話練習",
            "student_nationality": "美國 / United States",
            "student_level": "TOCFL B1",
        }
        data = {
            "user": self.tutor, "starts_on": date(2026, 7, 1), "ends_on": date(2026, 8, 31),
            "sections": [{"semester": self.semester, "rows": [detail_row] * 7, "subtotal": 7}],
            "total": 7, "generated_at": timezone.now(),
        }
        content = build_hours_pdf(
            data, version="detailed", detail_fields=("date", "nationality", "level", "hours"),
            program=self.ntnu_program, language="zh",
        )
        self.assertEqual(len(PdfReader(BytesIO(content)).pages), 2)

    def test_private_certificate_assets_are_embedded_when_provisioned(self):
        from pypdf import PdfReader

        data = {
            "user": self.tutor, "starts_on": date(2026, 7, 1), "ends_on": date(2026, 8, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        content = build_hours_pdf(data, version="summary", program=self.ntnu_program, language="zh")
        reader = PdfReader(BytesIO(content))
        page = reader.pages[0]
        font_names = {str(font.get_object().get("/BaseFont")) for font in page["/Resources"]["/Font"].values()}
        private_font_dir = Path(settings.BASE_DIR) / "assets/fonts"
        if (private_font_dir / "DFLiSongStd-W3.ttf").exists():
            self.assertTrue(any("DFLiSongStd-W3" in name for name in font_names))
        if (private_font_dir / "Helvetica Neue Condensed Bold.ttf").exists():
            self.assertTrue(any("HelveticaNeue-CondensedBold" in name for name in font_names))
        stamp_path = Path(settings.BASE_DIR) / "assets/certificates/CSL stamp.png"
        if stamp_path.exists():
            template_reader = PdfReader(
                Path(settings.BASE_DIR) / "tutoring/resources/certificate_templates/csl_template.pdf"
            )
            self.assertEqual(len(page.images), len(template_reader.pages[0].images) + 1)

    def test_certificate_shows_bilingual_name_when_both_names_present_in_both_languages(self):
        """Regression test: <b>陳安然</b> alone (no explicit font) renders blank inside an
        English-default Paragraph, because ReportLab resolves <b> through the paragraph's
        registered font family, and CertificateSerif's bold face has no CJK glyphs."""
        from pypdf import PdfReader

        self.tutor.name_zh = "陳安然"
        self.tutor.name_en = "Jamie Chen"
        self.tutor.save()
        data = {
            "user": self.tutor, "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        for language in ("zh", "en"):
            content = build_hours_pdf(data, version="summary", program=self.ntnu_program, language=language)
            text = PdfReader(BytesIO(content)).pages[0].extract_text()
            self.assertIn("陳安然", text)
            self.assertIn("Jamie Chen", text)

    def test_all_certificate_variants_put_opening_name_and_body_on_separate_lines(self):
        """Summary/detail and Chinese/English certificates share the same three-part lead:
        opening line, name plus student ID line, then the remaining certificate body."""
        from pypdf import PdfReader

        self.tutor.name_zh, self.tutor.name_en = "王小華", "Alex Wang"
        self.tutor.save()
        self.tutee.name_zh, self.tutee.name_en = "林安娜", "Anna Lin"
        self.tutee.save()
        base_data = {
            "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        cases = (
            (self.tutor, "zh", "本系碩士班學生"),
            (self.tutor, "en", "This certifies that graduate (Master's) student"),
            (self.tutee, "zh", "茲證明"),
            (self.tutee, "en", "This is to certify that"),
        )
        for user, language, opening in cases:
            for version in ("summary", "detailed"):
                data = {**base_data, "user": user}
                content = build_hours_pdf(
                    data,
                    version=version,
                    detail_fields=("date", "hours"),
                    program=self.ntnu_program,
                    language=language,
                )
                lines = [line.strip() for line in PdfReader(BytesIO(content)).pages[0].extract_text().splitlines()]
                opening_index = lines.index(opening)
                self.assertIn(user.username, lines[opening_index + 1])
                self.assertNotIn(user.username, lines[opening_index])
                self.assertNotIn(user.username, lines[opening_index + 2])

    def test_generic_chinese_period_has_no_indent_and_mentions_roc_only_once(self):
        from pypdf import PdfReader

        data = {
            "user": self.tutee,
            "starts_on": date(2026, 7, 1),
            "ends_on": date(2026, 8, 31),
            "sections": [],
            "total": 2,
            "generated_at": timezone.now(),
        }
        for version in ("summary", "detailed"):
            content = build_hours_pdf(
                data,
                version=version,
                detail_fields=("date", "hours"),
                program=self.ntnu_program,
                language="zh",
            )
            text = PdfReader(BytesIO(content)).pages[0].extract_text()
            compact = text.replace(" ", "").replace("\n", "")
            self.assertIn("於民國115年7月1日至8月31日期間", compact)
            self.assertNotIn("日至民國", compact)
            body_line = next(line for line in text.splitlines() if "於民國" in line.replace(" ", ""))
            self.assertTrue(body_line.startswith("於"))

    def test_certificate_single_name_shows_no_dangling_slash_in_both_languages(self):
        from pypdf import PdfReader

        self.tutor.name_zh = "陳安然"
        self.tutor.name_en = ""
        self.tutor.save()
        data = {
            "user": self.tutor, "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        for language in ("zh", "en"):
            content = build_hours_pdf(data, version="summary", program=self.ntnu_program, language=language)
            text = PdfReader(BytesIO(content)).pages[0].extract_text()
            self.assertIn("陳安然", text)
            self.assertNotIn("/", text)

    def test_missing_english_certificate_text_raises_clear_error(self):
        # The NTNU tutor branch has hardcoded body text and never reads
        # plan_name_en/activity_text_en (see test_ntnu_tutor_certificate_ignores_unused_plan_name_and_activity_fields
        # below), so only its title can trigger this validation.
        self.ntnu_program.tutor_certificate_title_en = ""
        self.ntnu_program.save()
        data = {
            "user": self.tutor, "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        with self.assertRaises(ValidationError) as ctx:
            build_hours_pdf(data, version="summary", program=self.ntnu_program, language="en")
        self.assertIn("請洽系辦設定", str(ctx.exception))

    def test_missing_english_certificate_text_raises_clear_error_for_generic_program(self):
        self.maryland_program.tutee_certificate_plan_name_en = ""
        self.maryland_program.save()
        maryland_tutee_roster = RosterEntry.objects.create(
            student_id="PP-TUTEE-MD4", name_zh="方案馬里蘭學生四", role=Role.TUTEE,
            education_level=EducationLevel.NOT_APPLICABLE, identity_category=IdentityCategory.INTERNATIONAL,
            program=self.maryland_program,
        )
        maryland_tutee = User.objects.create_user(
            username="PP-TUTEE-MD4", password="Password-2026", role=Role.TUTEE, roster_entry=maryland_tutee_roster
        )
        data = {
            "user": maryland_tutee, "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        with self.assertRaises(ValidationError) as ctx:
            build_hours_pdf(data, version="summary", program=self.maryland_program, language="en")
        self.assertIn("請洽系辦設定", str(ctx.exception))

    def test_missing_chinese_plan_name_raises_clear_error_for_generic_program(self):
        """Item 16: the Chinese path only ever validated title_zh before this fix, so a
        program with a title but a blank plan_name/activity_text silently rendered a
        certificate with an empty '「」' clause instead of failing clearly."""
        self.ntnu_program.tutee_certificate_plan_name = ""
        self.ntnu_program.save()
        data = {
            "user": self.tutee, "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        with self.assertRaises(ValidationError) as ctx:
            build_hours_pdf(data, version="summary", program=self.ntnu_program, language="zh")
        self.assertIn("請洽系辦設定", str(ctx.exception))

    def test_missing_chinese_activity_text_raises_clear_error_for_generic_program(self):
        self.ntnu_program.tutee_certificate_activity_text = ""
        self.ntnu_program.save()
        data = {
            "user": self.tutee, "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        with self.assertRaises(ValidationError) as ctx:
            build_hours_pdf(data, version="summary", program=self.ntnu_program, language="zh")
        self.assertIn("請洽系辦設定", str(ctx.exception))

    def test_ntnu_tutor_certificate_ignores_unused_plan_name_and_activity_fields(self):
        """Item 16: validation must match what each branch actually renders. The NTNU
        tutor branch has its own fixed-format sentence and never reads plan_name/
        activity_text, so leaving those blank must NOT block the download in either
        language (regression guard against over-strict validation)."""
        from pypdf import PdfReader

        self.ntnu_program.tutor_certificate_plan_name = ""
        self.ntnu_program.tutor_certificate_activity_text = ""
        self.ntnu_program.tutor_certificate_plan_name_en = ""
        self.ntnu_program.tutor_certificate_activity_text_en = ""
        self.ntnu_program.save()
        data = {
            "user": self.tutor, "starts_on": date(2026, 2, 1), "ends_on": date(2026, 7, 31),
            "sections": [], "total": 2, "generated_at": timezone.now(),
        }
        for language in ("zh", "en"):
            content = build_hours_pdf(data, version="summary", program=self.ntnu_program, language=language)
            text = PdfReader(BytesIO(content)).pages[0].extract_text()
            self.assertTrue(text.strip())

    def test_english_detailed_certificate_uses_english_table_headers_and_pagination_label(self):
        from pypdf import PdfReader

        past_date = timezone.localdate() - timedelta(days=12)
        self._make_verified_session(self.pairing, self.tutee, past_date)
        self.client.force_login(self.tutor)
        today = timezone.localdate()
        response = self.client.post(
            reverse("tutoring:download_hours"),
            {
                "mode": "range", "starts_on": today - timedelta(days=20), "ends_on": today,
                "version": "detailed", "program": self.ntnu_program.pk, "language": "en",
                "detail_fields": ["date", "nationality", "level", "hours"],
            },
        )
        self.assertEqual(response.status_code, 200)
        text = PdfReader(BytesIO(response.content)).pages[0].extract_text()
        self.assertIn("Hours Detail", text)
        self.assertIn("Page 1 of 1", text)
        self.assertIn("Nationality", text)
        self.assertIn("Chinese level", text)
        self.assertNotIn("輔導時數明細", text)


class ClassDocumentTests(MatchingFixtureTestCase):
    """Item 5: 合作計畫「上課文件」. class_documents_enabled is only True for MARYLAND in
    this first phase (see accounts/migrations/0014_enable_maryland_class_documents.py)."""

    def make_document(self, program, *, is_active=True, semester=None):
        return ClassDocument.objects.create(
            program=program, semester=semester, title_zh="教材", title_en="Course material",
            file=SimpleUploadedFile("material.pdf", b"%PDF-1.4 test", content_type="application/pdf"),
            is_active=is_active,
        )

    def test_maryland_tutee_and_bachelor_tutor_see_the_maryland_program(self):
        self.assertEqual(visible_class_document_programs(self.maryland), [self.maryland_program])
        self.assertEqual(visible_class_document_programs(self.maryland_tutor), [self.maryland_program])

    def test_ntnu_tutee_and_ordinary_tutor_see_nothing(self):
        """NTNU does not have class_documents_enabled in this first phase, so neither an
        NTNU tutee nor an ordinary tutor (who implicitly serves NTNU) sees any program."""
        self.assertEqual(visible_class_document_programs(self.tutee), [])
        self.assertEqual(visible_class_document_programs(self.tutor), [])

    def test_maryland_roster_tutor_without_bachelor_level_sees_nothing(self):
        """Mirrors tutor_can_serve_program()'s Maryland bachelor's restriction (item 4):
        a tutor on the Maryland roster but not at bachelor's level still can't serve
        Maryland, so they shouldn't see Maryland's class documents either."""
        graduate_maryland_tutor = self.make_tutor(
            "MARYTUTOR-GRAD", "馬里蘭碩士老師", "Maryland Grad Tutor",
            program=self.maryland_program, education_level=EducationLevel.MASTER,
        )
        self.assertEqual(visible_class_document_programs(graduate_maryland_tutor), [])

    def test_visible_class_documents_excludes_inactive_and_ineligible_programs(self):
        active_doc = self.make_document(self.maryland_program)
        self.make_document(self.maryland_program, is_active=False)
        self.make_document(self.ntnu_program)
        self.assertEqual(list(visible_class_documents(self.maryland)), [active_doc])
        self.assertEqual(list(visible_class_documents(self.tutee)), [])

    def test_class_documents_page_lists_only_eligible_active_documents(self):
        active_doc = self.make_document(self.maryland_program)
        self.make_document(self.ntnu_program)
        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:class_documents"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["documents"]), [active_doc])

    def test_menu_visibility_matches_eligibility(self):
        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertTrue(response.context["class_documents_visible"])

        self.client.force_login(self.tutee)
        response = self.client.get(reverse("accounts:dashboard"))
        self.assertFalse(response.context["class_documents_visible"])

    def test_eligible_user_can_download_and_audit_log_is_written(self):
        document = self.make_document(self.maryland_program)
        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:download_class_document", args=[document.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("material.pdf", response["Content-Disposition"])
        self.assertEqual(response["Cache-Control"], "private, no-store")
        self.assertEqual(response["X-Content-Type-Options"], "nosniff")
        log = AuditLog.objects.get(event_type="CLASS_DOCUMENT_DOWNLOADED")
        self.assertEqual(log.actor, self.maryland)
        self.assertEqual(log.metadata["program"], "MARYLAND")

    def test_stored_filename_is_randomized_but_original_name_is_kept(self):
        """Batch 3 item 5, applied consistently to the third file type: Admin-uploaded
        class documents also get a UUID-based stored filename, with the original name
        preserved via ClassDocument.original_filename for display/Content-Disposition."""
        document = self.make_document(self.maryland_program)
        self.assertEqual(document.filename, "material.pdf")
        self.assertEqual(document.original_filename, "material.pdf")
        self.assertNotIn("material", document.file.name)

    def test_ineligible_user_cannot_download(self):
        document = self.make_document(self.maryland_program)
        self.client.force_login(self.tutee)
        response = self.client.get(reverse("accounts:download_class_document", args=[document.pk]))
        self.assertEqual(response.status_code, 404)

    def test_inactive_document_cannot_be_downloaded_even_by_eligible_user(self):
        document = self.make_document(self.maryland_program, is_active=False)
        self.client.force_login(self.maryland)
        response = self.client.get(reverse("accounts:download_class_document", args=[document.pk]))
        self.assertEqual(response.status_code, 404)


READ_ONLY_ADMIN_CLASSES = [
    (PairingAdmin, Pairing),
    (MatchingInvitationAdmin, MatchingInvitation),
    (ClassSessionAdmin, ClassSession),
    (AttendanceAdmin, Attendance),
    (ClassRecordAdmin, ClassRecord),
    (ClassConfirmationAdmin, ClassConfirmation),
    (MakeupReviewAdmin, MakeupReview),
    (PairingReleaseRequestAdmin, PairingReleaseRequest),
]


class CoreModelAdminReadOnlyTests(TestCase):
    """Batch 4 item 3 (docs/VULNERABILITY_SCAN_IMPROVEMENTS.md): the 8 core business
    models must be read-only in Django Admin so an Admin login can't add/change/delete
    records in a way that bypasses tutoring/services.py's quota, status, and locking
    rules. Checked two ways: directly against each ModelAdmin's permission methods (fast,
    covers all 8 without needing a fixture per model) below, then end-to-end through the
    test client for one representative model (PairingAdminViewTests) to confirm Django
    Admin actually enforces it at the HTTP layer, not just that the method returns False
    in isolation."""

    def setUp(self):
        self.admin_user = User.objects.create_superuser(username="ADMIN-RO-TEST", password="Admin-password-2026")
        self.request = RequestFactory().get("/system-admin/")
        self.request.user = self.admin_user

    def test_add_change_delete_all_denied_even_for_a_real_superuser(self):
        for admin_class, model in READ_ONLY_ADMIN_CLASSES:
            with self.subTest(model=model.__name__):
                instance = admin_class(model, django_admin.site)
                self.assertFalse(instance.has_add_permission(self.request))
                self.assertFalse(instance.has_change_permission(self.request))
                self.assertFalse(instance.has_change_permission(self.request, obj=object()))
                self.assertFalse(instance.has_delete_permission(self.request))
                self.assertFalse(instance.has_delete_permission(self.request, obj=object()))


class PairingAdminViewTests(TestCase):
    """End-to-end confirmation, for one representative model, that Django Admin itself
    enforces the ReadOnlyAdminMixin permissions (view/changelist still work, add/change/
    delete are blocked) rather than just the permission methods returning False."""

    def setUp(self):
        self.admin = User.objects.create_superuser(username="ADMIN-PAIRING-RO", password="Admin-password-2026")
        semester = Semester.objects.create(
            name_zh="唯讀測試學期", name_en="Read-only test semester",
            starts_on=date(2026, 1, 1), ends_on=date(2026, 6, 30), is_active=True,
        )
        tutor = User.objects.create_user(username="RO-TUTOR", password="Test-password-2026", role=Role.TUTOR)
        tutee = User.objects.create_user(username="RO-TUTEE", password="Test-password-2026", role=Role.TUTEE)
        self.pairing = Pairing.objects.create(semester=semester, tutor=tutor, tutee=tutee)
        self.client.force_login(self.admin)

    def test_changelist_and_detail_are_still_viewable(self):
        response = self.client.get(reverse("admin:tutoring_pairing_changelist"))
        self.assertEqual(response.status_code, 200)
        response = self.client.get(reverse("admin:tutoring_pairing_change", args=[self.pairing.pk]))
        self.assertEqual(response.status_code, 200)

    def test_add_view_is_forbidden(self):
        response = self.client.get(reverse("admin:tutoring_pairing_add"))
        self.assertEqual(response.status_code, 403)

    def test_change_post_is_forbidden(self):
        response = self.client.post(reverse("admin:tutoring_pairing_change", args=[self.pairing.pk]), {})
        self.assertEqual(response.status_code, 403)
        self.pairing.refresh_from_db()
        self.assertEqual(self.pairing.status, PairingStatus.ACTIVE)

    def test_delete_view_is_forbidden(self):
        response = self.client.post(reverse("admin:tutoring_pairing_delete", args=[self.pairing.pk]), {"post": "yes"})
        self.assertEqual(response.status_code, 403)
        self.assertTrue(Pairing.objects.filter(pk=self.pairing.pk).exists())


class SpreadsheetInjectionTests(TestCase):
    """Batch 6 item 2 (docs/VULNERABILITY_SCAN_IMPROVEMENTS.md): CSV/XLSX exports must
    neutralize values that would be interpreted as spreadsheet formulas."""

    def setUp(self):
        self.tutor = User.objects.create_user(
            username="INJECT-TUTOR", password="Test-password-2026", role=Role.TUTOR,
            name_zh='=HYPERLINK("https://evil.example")', name_en="+SUM(1,2)",
        )

    def test_csv_export_escapes_formula_prefixed_values(self):
        content = build_export_csv([self.tutor])
        text = content.decode("utf-8-sig")
        self.assertIn("'=HYPERLINK", text)
        self.assertIn("'+SUM", text)

    def test_xlsx_export_stores_formula_prefixed_values_as_literal_text(self):
        content = build_excel_xlsx([self.tutor])
        workbook = openpyxl.load_workbook(BytesIO(content))
        data_rows = list(workbook.active.iter_rows(min_row=2))
        cell_values = {cell.value for row in data_rows for cell in row}
        self.assertIn('\'=HYPERLINK("https://evil.example")', cell_values)
        self.assertIn("'+SUM(1,2)", cell_values)
        # openpyxl only classifies a cell as a formula (data_type "f") when the raw value
        # it was given starts with "="; the leading apostrophe must prevent that.
        for row in data_rows:
            for cell in row:
                self.assertNotEqual(cell.data_type, "f")

    def test_ordinary_names_are_not_altered(self):
        ordinary = User.objects.create_user(
            username="ORDINARY-TUTOR", password="Test-password-2026", role=Role.TUTOR, name_zh="王小明",
        )
        content = build_export_csv([ordinary])
        text = content.decode("utf-8-sig")
        self.assertIn("王小明", text)
        self.assertNotIn("'王小明", text)


def minimal_png_bytes(size=(10, 10)):
    from PIL import Image

    buffer = BytesIO()
    Image.new("RGB", size, color=(255, 0, 0)).save(buffer, format="PNG")
    return buffer.getvalue()


def minimal_docx_bytes():
    """A ZIP archive with the one manifest entry validate_class_document_file actually
    checks for — not a fully valid Word document, since the validator doesn't parse the
    document body, only that it's a real ZIP with that entry."""
    import zipfile

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
    return buffer.getvalue()


class UploadContentValidationTests(TestCase):
    """Batch 6 item 1 (docs/VULNERABILITY_SCAN_IMPROVEMENTS.md): uploads must be validated
    by actual content, not just extension/size, so a renamed HTML/script file can't pass
    as a PDF/JPG/PNG/Office document."""

    def upload(self, name, content, content_type):
        return SimpleUploadedFile(name, content, content_type=content_type)

    def test_fake_pdf_content_is_rejected(self):
        upload = self.upload("fake.pdf", b"not actually a pdf", "application/pdf")
        with self.assertRaises(ValidationError):
            validate_qualification_file(upload)

    def test_real_pdf_content_is_accepted(self):
        upload = self.upload("real.pdf", minimal_pdf_bytes(), "application/pdf")
        validate_qualification_file(upload)

    def test_pdf_page_count_over_the_limit_is_rejected(self):
        from pypdf import PdfWriter

        writer = PdfWriter()
        for _ in range(3):
            writer.add_blank_page(width=72, height=72)
        buffer = BytesIO()
        writer.write(buffer)
        upload = self.upload("many-pages.pdf", buffer.getvalue(), "application/pdf")
        with self.assertRaises(ValidationError), patch("tutoring.models.MAX_PDF_PAGES", 2):
            validate_qualification_file(upload)

    def test_fake_image_content_is_rejected(self):
        upload = self.upload("fake.png", b"not actually a png", "image/png")
        with self.assertRaises(ValidationError):
            validate_qualification_file(upload)

    def test_real_image_content_is_accepted(self):
        upload = self.upload("real.png", minimal_png_bytes(), "image/png")
        validate_qualification_file(upload)

    def test_oversized_image_dimensions_are_rejected(self):
        upload = self.upload("huge.png", minimal_png_bytes(size=(6001, 2)), "image/png")
        with self.assertRaises(ValidationError):
            validate_class_record_attachment(upload)

    def test_fake_docx_content_is_rejected(self):
        upload = self.upload("fake.docx", b"not actually a docx", "application/octet-stream")
        with self.assertRaises(ValidationError):
            validate_class_document_file(upload)

    def test_real_docx_content_is_accepted(self):
        upload = self.upload("real.docx", minimal_docx_bytes(), "application/octet-stream")
        validate_class_document_file(upload)

    def test_fake_legacy_doc_content_is_rejected(self):
        upload = self.upload("fake.doc", b"not actually a doc", "application/octet-stream")
        with self.assertRaises(ValidationError):
            validate_class_document_file(upload)

    def test_real_legacy_doc_header_is_accepted(self):
        header = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        upload = self.upload("real.doc", header + b"\x00" * 100, "application/octet-stream")
        validate_class_document_file(upload)
