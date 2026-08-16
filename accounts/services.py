import csv
import io
import re
from dataclasses import dataclass, field

import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction

from .models import EducationLevel, IdentityCategory, PartnerProgram, Role, RosterEntry

ROSTER_IMPORT_COLUMNS = [
    "student_id",
    "name_zh",
    "name_en",
    "role",
    "education_level",
    "identity_category",
    "program_code",
    "is_enabled",
]

_TRUE_VALUES = {"true", "1", "yes", "y", "是", "true "}
_FALSE_VALUES = {"false", "0", "no", "n", "否", ""}


class RosterImportFileError(Exception):
    """檔案本身無法解析（格式錯誤、空檔案、副檔名不支援）。"""


@dataclass
class RosterImportResult:
    created_count: int = 0
    created_ids: list = field(default_factory=list)
    updated_count: int = 0
    updated_ids: list = field(default_factory=list)
    skipped_existing_ids: list = field(default_factory=list)
    skipped_invalid: list = field(default_factory=list)
    errors: list = field(default_factory=list)


def _read_csv_rows(uploaded_file):
    raw = uploaded_file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))
    if reader.fieldnames is None:
        raise RosterImportFileError("檔案是空的。 / The file is empty.")
    return list(reader)


def _read_xlsx_rows(uploaded_file):
    workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise RosterImportFileError("檔案是空的。 / The file is empty.")
    header = [str(cell).strip() if cell is not None else "" for cell in header]
    rows = []
    for raw_row in rows_iter:
        if raw_row is None or all(cell is None for cell in raw_row):
            continue
        row = {header[i]: raw_row[i] for i in range(len(header)) if i < len(raw_row)}
        rows.append(row)
    return rows


def _parse_bool(raw_value, row_num, field_name, errors):
    value = "" if raw_value is None else str(raw_value).strip().lower()
    if value in _TRUE_VALUES and value != "":
        return True
    if value in _FALSE_VALUES:
        return False
    errors.append(f"第 {row_num} 列：{field_name} 不是有效的布林值「{raw_value}」。 / Row {row_num}: {field_name} is not a valid boolean.")
    return None


def _clean_str(raw_value):
    if raw_value is None:
        return ""
    return str(raw_value).strip()


def parse_roster_import_rows(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return _read_csv_rows(uploaded_file)
    if name.endswith(".xlsx"):
        return _read_xlsx_rows(uploaded_file)
    raise RosterImportFileError("僅支援 .csv 或 .xlsx 檔案。 / Only .csv or .xlsx files are supported.")


def validate_roster_import_rows(raw_rows):
    errors = []
    entries = []
    skipped_existing_ids = []
    seen_student_ids = set()
    existing_student_ids = set(RosterEntry.objects.values_list("student_id", flat=True))
    programs_by_code = {program.code: program for program in PartnerProgram.objects.all()}

    for index, raw_row in enumerate(raw_rows, start=2):  # row 1 is the header
        student_id = _clean_str(raw_row.get("student_id")).upper()
        name_zh = _clean_str(raw_row.get("name_zh"))
        name_en = _clean_str(raw_row.get("name_en"))
        role = _clean_str(raw_row.get("role")).upper()
        education_level = _clean_str(raw_row.get("education_level")).upper() or EducationLevel.NOT_APPLICABLE
        identity_category = _clean_str(raw_row.get("identity_category")).upper()
        program_code = _clean_str(raw_row.get("program_code")).upper()
        is_enabled = _parse_bool(raw_row.get("is_enabled", "true"), index, "is_enabled", errors)

        if not student_id:
            errors.append(f"第 {index} 列：學號為必填欄位。 / Row {index}: student_id is required.")
            continue
        if not name_zh:
            errors.append(f"第 {index} 列：中文姓名為必填欄位。 / Row {index}: name_zh is required.")
        if role not in {Role.TUTOR, Role.TUTEE}:
            errors.append(f"第 {index} 列：身分「{role}」不是合法值（TUTOR / TUTEE）。 / Row {index}: role must be TUTOR or TUTEE.")
        if education_level not in EducationLevel.values:
            errors.append(f"第 {index} 列：學制「{education_level}」不是合法值。 / Row {index}: invalid education_level.")
        if identity_category not in IdentityCategory.values:
            errors.append(f"第 {index} 列：學生類別「{identity_category}」不是合法值。 / Row {index}: invalid identity_category.")
        program = programs_by_code.get(program_code) if program_code and program_code != "NA" else None
        if program_code and program_code != "NA" and program is None:
            errors.append(f"第 {index} 列：所屬計畫代碼「{program_code}」不存在。 / Row {index}: unknown program_code.")
        if role == Role.TUTEE and program is None:
            errors.append(f"第 {index} 列：Tutee 必須設定所屬計畫。 / Row {index}: program_code is required for tutees.")

        if student_id in seen_student_ids:
            continue  # already handled earlier in this same file; keep the first occurrence
        if student_id in existing_student_ids:
            skipped_existing_ids.append(student_id)
            seen_student_ids.add(student_id)
            continue
        seen_student_ids.add(student_id)

        row_had_error = any(err.startswith(f"第 {index} 列") for err in errors)
        if row_had_error:
            continue

        entry = RosterEntry(
            student_id=student_id,
            name_zh=name_zh,
            name_en=name_en,
            role=role,
            education_level=education_level,
            identity_category=identity_category,
            program=program,
            is_enabled=is_enabled if is_enabled is not None else True,
        )
        try:
            entry.full_clean(exclude=["id"])
        except ValidationError as exc:
            for messages_list in exc.message_dict.values():
                for message in messages_list:
                    errors.append(f"第 {index} 列：{message} / Row {index}: {message}")
            continue
        entries.append(entry)

    return entries, errors, skipped_existing_ids


_TEMPLATE_SAMPLE_ROWS = [
    ["S10112345", "王小明", "Wang Xiao-Ming", "TUTOR", "MASTER", "LOCAL", "NA", "TRUE"],
    ["S20223456", "陳小美", "Chen Xiao-Mei", "TUTEE", "NA", "INTERNATIONAL", "NTNU", "TRUE"],
]


def roster_template_csv_bytes():
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(ROSTER_IMPORT_COLUMNS)
    writer.writerows(_TEMPLATE_SAMPLE_ROWS)
    return buffer.getvalue().encode("utf-8-sig")


def roster_template_xlsx_bytes():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(ROSTER_IMPORT_COLUMNS)
    for row in _TEMPLATE_SAMPLE_ROWS:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def import_roster_entries(uploaded_file):
    raw_rows = parse_roster_import_rows(uploaded_file)
    if not raw_rows:
        return RosterImportResult(errors=["檔案沒有任何資料列。 / The file has no data rows."])

    entries, errors, skipped_existing_ids = validate_roster_import_rows(raw_rows)
    if errors:
        return RosterImportResult(errors=errors)

    with transaction.atomic():
        for entry in entries:
            entry.save()

    return RosterImportResult(
        created_count=len(entries),
        created_ids=[entry.student_id for entry in entries],
        skipped_existing_ids=skipped_existing_ids,
    )


_STUDENT_ID_PATTERN = re.compile(r"^[A-Za-z0-9]{4,24}$")

_STUDENT_ID_HEADERS = {"學號", "學生學號", "studentid", "student_id"}
_IDENTITY_IMPORT_ALIASES = {
    "local": IdentityCategory.LOCAL,
    "domesticstudent": IdentityCategory.LOCAL,
    "本地生": IdentityCategory.LOCAL,
    "本國生": IdentityCategory.LOCAL,
    "overseas": IdentityCategory.OVERSEAS,
    "overseaschinesestudent": IdentityCategory.OVERSEAS,
    "僑生": IdentityCategory.OVERSEAS,
    "hongkongmacao": IdentityCategory.HONG_KONG_MACAO,
    "hongkongandmacaostudent": IdentityCategory.HONG_KONG_MACAO,
    "港澳生": IdentityCategory.HONG_KONG_MACAO,
    "mainland": IdentityCategory.MAINLAND,
    "mainlandchinesestudent": IdentityCategory.MAINLAND,
    "陸生": IdentityCategory.MAINLAND,
    "international": IdentityCategory.INTERNATIONAL,
    "internationalstudent": IdentityCategory.INTERNATIONAL,
    "foreignstudent": IdentityCategory.INTERNATIONAL,
    "外籍生": IdentityCategory.INTERNATIONAL,
    "外國學生": IdentityCategory.INTERNATIONAL,
    "國際學生": IdentityCategory.INTERNATIONAL,
    # Maryland is represented by PartnerProgram rather than a fifth model-level identity;
    # its registration selector derives MARYLAND from role + program.
    "marylandstudent": IdentityCategory.INTERNATIONAL,
    "馬里蘭學生": IdentityCategory.INTERNATIONAL,
}


def _compact_import_label(raw_value):
    return re.sub(r"[\s/_\-]+", "", _clean_str(raw_value)).lower()


def _parse_quick_identity(raw_value):
    value = _clean_str(raw_value)
    if not value:
        return ""
    compact = _compact_import_label(value)
    if compact in _IDENTITY_IMPORT_ALIASES:
        return _IDENTITY_IMPORT_ALIASES[compact]
    upper_value = value.upper()
    return upper_value if upper_value in IdentityCategory.values else None


def _read_quick_roster_rows(uploaded_file):
    """Read student ID and optional identity from the first two columns.

    Source lists are often messy (a title row, a Chinese header row, stray
    whitespace) rather than a clean export, so this reads every row rather than
    assuming a specific header. The second column remains optional for backwards
    compatibility with historical ID-only lists.
    """
    name = uploaded_file.name.lower()
    values = []
    if name.endswith(".csv"):
        raw = uploaded_file.read().decode("utf-8-sig")
        for row_num, row in enumerate(csv.reader(io.StringIO(raw)), start=1):
            if row and row[0] is not None and str(row[0]).strip():
                identity = row[1] if len(row) > 1 else ""
                values.append((row_num, str(row[0]).strip(), identity))
    elif name.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cell = row[0] if row else None
            if cell is not None and str(cell).strip():
                identity = row[1] if len(row) > 1 else ""
                values.append((row_num, str(cell).strip(), identity))
    else:
        raise RosterImportFileError("僅支援 .csv 或 .xlsx 檔案。 / Only .csv or .xlsx files are supported.")
    return values


def import_roster_ids(uploaded_file, *, role, program=None):
    """Quick import: student IDs plus an optional identity-category column.

    Role/program still come from the Admin card selected for upload. Identity is
    read from column two when present; re-uploading may fill a previously blank
    identity but never overwrites an existing non-blank administrative value.
    """
    raw_values = _read_quick_roster_rows(uploaded_file)
    if not raw_values:
        return RosterImportResult(errors=["檔案沒有任何資料列。 / The file has no data rows."])

    candidates = {}
    skipped_invalid = []
    for row_num, raw_value, raw_identity in raw_values:
        candidate = raw_value.strip().upper()
        if _compact_import_label(raw_value) in _STUDENT_ID_HEADERS:
            continue
        if not _STUDENT_ID_PATTERN.match(candidate):
            skipped_invalid.append(f"第 {row_num} 列：「{raw_value}」不是有效學號格式，已略過。 / Row {row_num}: not a valid student ID, skipped.")
            continue
        identity_category = _parse_quick_identity(raw_identity)
        if identity_category is None:
            skipped_invalid.append(
                f"第 {row_num} 列：無法識別身分別「{raw_identity}」，已略過。 / "
                f"Row {row_num}: unknown identity category, skipped."
            )
            continue
        if candidate not in candidates or (not candidates[candidate] and identity_category):
            candidates[candidate] = identity_category

    existing_entries = RosterEntry.objects.in_bulk(candidates, field_name="student_id")
    new_ids = [student_id for student_id in candidates if student_id not in existing_entries]
    updated_ids = []
    skipped_existing_ids = []

    with transaction.atomic():
        for student_id in new_ids:
            entry = RosterEntry(
                student_id=student_id,
                role=role,
                program=program,
                identity_category=candidates[student_id],
            )
            entry.full_clean(exclude=["id"])
            entry.save()
        for student_id, entry in existing_entries.items():
            incoming_identity = candidates[student_id]
            if incoming_identity and not entry.identity_category:
                entry.identity_category = incoming_identity
                entry.full_clean(exclude=["id"])
                entry.save(update_fields=["identity_category", "updated_at"])
                updated_ids.append(student_id)
            else:
                skipped_existing_ids.append(student_id)
                if incoming_identity and entry.identity_category != incoming_identity:
                    skipped_invalid.append(
                        f"學號 {student_id} 已有不同身分別，保留系統原資料。 / "
                        f"Student ID {student_id} already has a different identity; existing data was kept."
                    )

    return RosterImportResult(
        created_count=len(new_ids),
        created_ids=new_ids,
        updated_count=len(updated_ids),
        updated_ids=sorted(updated_ids),
        skipped_existing_ids=sorted(skipped_existing_ids),
        skipped_invalid=skipped_invalid,
    )
